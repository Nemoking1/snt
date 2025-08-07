from openpyxl import load_workbook, Workbook
from sinotrans.core.Rule import Rule
from typing import get_type_hints
import os

class FileProcessor:
    @staticmethod
    def ensure_directories_exist(directories):
        """确保所有必要的目录存在，如果不存在则创建它们"""
        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
    @staticmethod
    def create_newfile_by_template(template_file_name, target_file_name, column_names = None):
            """创建新文件，并复制模板表头"""
            # 加载模板并获取表头（第一行数据）
            header_row = next(load_workbook(template_file_name).active.iter_rows(max_row=1, values_only=True))  # 提取第一行数据
            # 添加CRD列
            header_row = list(header_row)
            if not column_names:
                header_row.extend(column_names)

            # 创建新工作簿并写入表头
            new_wb = Workbook()
            new_sheet = new_wb.active
            new_sheet.append(header_row)
            
            new_wb.save(target_file_name)
            return new_wb
    @staticmethod
    def parse_mapping(file_name, splitter, prefix_separator, condition_separator, key_value_separator):
        """解析键值对格式的映射规则"""
        mapping = {}
        with open(file_name, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or splitter not in line:
                    continue
                key, value = line.split(splitter, 1)
                rule = FileProcessor.parse_rule(value, prefix_separator, condition_separator, key_value_separator)
                mapping[key.strip()] = rule
        return mapping
    @staticmethod
    def parse_mapping_list(file_name, splitter, prefix_separator, condition_separator, key_value_separator):
        """解析键值对格式的映射规则"""
        mapping = {}
        with open(file_name, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or splitter not in line:
                    continue
                key, value = line.split(splitter, 1)
                if key not in mapping:
                    mapping[key] = []
                rule = FileProcessor.parse_rule(value, prefix_separator, condition_separator, key_value_separator)
                mapping[key.strip()].append(rule)
        return mapping
    @staticmethod
    def parse_rule(value, prefix_separator, condition_separator, key_value_separator):
        rule = Rule(field_name=value.split('|')[0].strip())
        type_hints = get_type_hints(type(rule))
        # 解析参数键值对
        if prefix_separator in value:
            # 先取条件字符串
            params_str = value.split(prefix_separator, 1)[1]
            # 遍历条件，解析键值对
            for param in params_str.split(condition_separator):
                # param不要去掉末尾空格，不然分隔符为空格的时候就会被误删！
                param = param
                if key_value_separator not in param:
                    continue
                # 分割当前条件，提取键值对
                k, v = param.split(key_value_separator, 1)
                k = k.strip().lower()
                if hasattr(rule, k):
                    attr_type = type_hints[k]
                    try:
                        converted_v = attr_type(v)  # 尝试转换
                        setattr(rule, k, converted_v)
                    except (TypeError, ValueError) as e:
                        raise ValueError(f"❌ 参数 {k} 的值 {v} 无法转换为正确类型: {e}")
        return rule