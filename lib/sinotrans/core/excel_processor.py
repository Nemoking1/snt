from sinotrans.utils.progress_manager import ExcelProgressTracker
from sinotrans.utils.global_thread_pool import GlobalThreadPool
from sinotrans.utils.logger import Logger
from sinotrans.core.rule import Rule
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from deprecated import deprecated
from zipfile import BadZipFile
from pathlib import Path
import concurrent.futures
import pandas as pd
import warnings
import traceback
import time
import xlrd
import os

class ExcelProcessor:
    DEFAULT_SPLITTER = '#'
    """
    Excel映射关系工具处理类
    """
    def __init__(self):
        return
        # self.mapping = mapping, mapping: Dict[str, Any], excel_path:str
        # self.excel_path = excel_path
    @staticmethod
    def sort_generated_rows(mapped_rows, headers):
        """根据模板列顺序，构建排好序的新行列表"""
        try:
            ordered_rows = [] 
            for row in mapped_rows:
                # 按模板列顺序构建数据
                ordered_row = [
                    row.get(header, None)
                    for header in headers
                ]
                ordered_rows.append(ordered_row)
            return ordered_rows
        except Exception as e:
            Logger.error(f"❌ 构建新行数据失败: {str(e)}")
            raise
    @staticmethod
    def fixed_mapping(fixed_mapping):
        """根据固定字段映射，填充到对应列"""
        mapped_row = {}
        try:
            for dest_col, rule in fixed_mapping.items():
                mapped_row[dest_col] = rule.field_name
            return mapped_row
        except Exception as e:
            Logger.error(f"❌ fixed_mapping映射失败: {str(e)}")
            raise
    @staticmethod
    def column_mapping(row, column_mapping):
        """
        根据原行数据，结合字段映射文件，生成新行数据
        对当前输入进行map映射，映射到dest_name，返回字典{dest_name:value,......}
        """
        mapped_row = {}
        try:
            for src_col, rules in column_mapping.items():
                raw_value = row.get(src_col, None)
                for rule in rules:
                    if raw_value:
                        mapped_row[rule.field_name] = rule.map_action(raw_value)
                    if rule.considerEmpty:
                        mapped_row[rule.field_name] = raw_value
            return mapped_row
        except Exception as e:
            Logger.error(f"❌ mapping映射失败: {str(e)}")
            raise
    @staticmethod
    def email_mapping(row, key_field, global_po_mapping, email_mapping):
        """
        根据email映射和 global_po_mapping, 生成新行数据
        """
        mapped_row = {}
        try:
            # email_po值映射, 处理从邮件中获取的值——global_po_mapping, 合并邮件字段数据
            key = str(row[key_field])
            if key in global_po_mapping:
                for rules in email_mapping.values():
                    for rule in rules:
                        raw_value = global_po_mapping[key].get(rule.field_name)
                        mapped_row[rule.field_name] = rule.map_action(raw_value)
            return mapped_row
        except Exception as e:
            Logger.error(f"❌ email_mapping映射失败: {str(e)}")
            raise
    @staticmethod
    def content_mapping(key, map_content, map):
        """
        根据map和 map_content键值对合集, 生成新行数据
        """
        mapped_row = {}
        try:
            content = map_content.get(key)
            for key in content.keys():
                for rules in map.get(key):
                    for rule in rules:
                        raw_value = content.get(key)
                        mapped_row[key] = rule.map_action(raw_value)
            return mapped_row
        except Exception as e:
            Logger.error(f"❌ mapping映射失败: {str(e)}")
            raise

    @staticmethod
    def excel_row_generator_skipping(rs_input, file_name, progress=None, 
                        required_columns=None, desc=None, strict_flag=True):
        """优化后的行数据生成器（支持连续空行1000行提前终止）"""
        Logger.debug(f"📋 开始解析文件 {file_name}（共{rs_input.max_row}行）")
        headers = [cell.value for cell in rs_input[1]]
        
        # 预生成必填列检查器
        required_check = None
        if required_columns:
            required_check = {
                col: headers.index(col) 
                for col in required_columns 
                if col in headers
            }

        # 初始化连续空行计数器
        MAX_CONSECUTIVE_EMPTY = 1000  # 最大允许连续空行数
        empty_counter = 0
        
        for row_idx, row in enumerate(rs_input.iter_rows(min_row=2), start=2):
            # 空行检测
            row_data = {
                headers[idx]: cell.value.strip() if isinstance(cell.value, str) else cell.value
                for idx, cell in enumerate(row)
            }
            
            if all(v in (None, "") for v in row_data.values()):
                empty_counter += 1
                if empty_counter >= MAX_CONSECUTIVE_EMPTY:
                    Logger.debug(f"⏹ 检测到连续{empty_counter}行空行，提前终止读取（从第{row_idx}行起）")
                    break  # 直接终止循环
                continue
            else:
                empty_counter = 0  # 遇到非空行时重置计数器

            # 必填列校验逻辑（原有逻辑保持不变）
            if required_check:
                missing_cols = [
                    col for col, idx in required_check.items() 
                    if row_data.get(col) in (None, "")
                ]
                if strict_flag and missing_cols:
                    Logger.debug(f"严格模式跳过（第{row_idx}行），缺失字段：{', '.join(missing_cols)}")
                    continue
                elif not strict_flag and len(missing_cols) == len(required_check):
                    Logger.debug(f"宽松模式跳过（第{row_idx}行），全部必填字段缺失")
                    continue

            yield row_data

        Logger.debug(f"✅ 文件解析完成，实际处理到第{row_idx}行")
    @staticmethod
    def excel_row_generator(rs_input, file_name, progress=None, required_columns=None, desc=None, strict_flag=True):
        """
        带严格模式控制的行数据生成器
        参数：
        rs_input: xlrd.sheet.Sheet对象
        file_name: str, 文件名
        progress: ExcelProgressTracker对象, 进度管理器，可选
        required_columns: list, 必填列
        desc: str, 进度描述
        strict_flag: bool, 严格模式是否开启
        """
        Logger.debug(f"📋 开始解析文件{file_name}")
        headers = [cell.value for cell in rs_input[1]]
        
        # 预处理必填列索引
        required_indices = []
        if required_columns:
            required_indices = [
                headers.index(col) 
                for col in required_columns 
                if col in headers
            ]
        if progress:
            progress.init_main_progress(desc=desc, total=rs_input.max_row - 1)
        for row_idx, row in enumerate(rs_input.iter_rows(min_row=2), start=2):
            if progress:
                progress.update()  # 保持进度更新
            
            try:
                row_data = {
                    headers[idx]: cell.value.strip() if isinstance(cell.value, str) else cell.value
                    for idx, cell in enumerate(row)
                }

                # 空行检测
                if all(v in (None, "") for v in row_data.values()):
                    Logger.debug(f"- 跳过全空行（第{row_idx}行）")
                    continue

                # 必填列校验逻辑
                if required_indices:
                    missing_cols = [
                        headers[idx] 
                        for idx in required_indices 
                        if row_data.get(headers[idx]) in (None, "")
                    ]

                    # 严格模式：存在缺失即跳过
                    if strict_flag:
                        if missing_cols:
                            Logger.debug(f"严格模式跳过（第{row_idx}行），缺失字段：{', '.join(missing_cols)}")
                            continue
                    # 宽松模式：仅当全部缺失时跳过
                    else:
                        if len(missing_cols) == len(required_indices):
                            Logger.debug(f"宽松模式跳过（第{row_idx}行），全部必填字段缺失")
                            continue
                        elif missing_cols:
                            Logger.debug(f"宽松模式保留（第{row_idx}行），部分缺失字段：{', '.join(missing_cols)}")

                yield row_data

            except Exception as e:
                Logger.error(f"❌ 第{row_idx}行数据解析失败: {str(e)}")
                continue
    def _process_common(self, file_path: str, worksheet, map: Dict[str, List[Rule]], is_xlsx: bool) -> Tuple[str, dict]:
        """文件名，{目的段名：目的段值}"""
        filename_no_ext = os.path.splitext(os.path.basename(file_path))[0]
        dict_map = {}

        # 统一获取行列范围（兼容不同库的索引方式）
        if is_xlsx:
            row_range = range(1, worksheet.max_row + 1)  # openpyxl从1开始到max_row
            col_range = range(1, worksheet.max_column + 1)
        else:
            row_range = range(worksheet.nrows)  # xlrd从0开始到nrows-1
            col_range = range(worksheet.ncols)

        # 双重循环遍历所有单元格
        for row_idx in row_range:
            for col_idx in col_range:
                # 统一获取单元格值
                try:
                    cell_value = worksheet.cell(row_idx, col_idx).value if is_xlsx \
                        else worksheet.cell_value(row_idx, col_idx)
                except Exception as e:
                    Logger.error(f"读取单元格错误 @ 行{row_idx} 列{col_idx}: {str(e)}")
                    continue

                if cell_value in map:
                    for rule in map[cell_value]:
                        values = []
                        offset = 0
                        try:
                            # 确定目标位置和偏移方向
                            if rule.dir == "row":
                                target_row = row_idx + 1
                                target_col = col_idx
                                max_limit = worksheet.max_row if is_xlsx else worksheet.nrows
                            elif rule.dir == "column":
                                target_row = row_idx
                                target_col = col_idx + 1
                                max_limit = worksheet.max_column if is_xlsx else worksheet.ncols
                            # readingMode优先级较高
                            if rule.readingMode:
                                if rule.readingMode == "readUntilBlank":
                                    while True:
                                        # 计算当前读取位置
                                        pos = (
                                            target_row + (offset if rule.dir == "row" else 0),
                                            target_col + (offset if rule.dir == "column" else 0)
                                        )
                                        
                                        # 检查是否超出工作表范围
                                        if (rule.dir == "row" and pos[0] >= max_limit) or \
                                        (rule.dir == "column" and pos[1] >= max_limit):
                                            break
                                        
                                        # 获取单元格值
                                        cell_value = worksheet.cell(*pos).value if is_xlsx else worksheet.cell_value(*pos)
                                        if cell_value is None or cell_value == "":
                                            break
                                        
                                        values.append(str(cell_value))
                                        offset += 1
                                    
                                    target_value = self.DEFAULT_SPLITTER.join(values) if values else ""
                            elif rule.count:
                                for i in range(rule.count):
                                    pos = (target_row + (offset if rule.dir == "row" else 0), 
                                        target_col + (offset if rule.dir == "column" else 0))
                                    
                                    if (rule.dir == "row" and pos[0] >= max_limit) or \
                                    (rule.dir == "column" and pos[1] >= max_limit):
                                        raise IndexError("偏移超出工作表范围")
                                    
                                    cell_value = worksheet.cell(*pos).value if is_xlsx else worksheet.cell_value(*pos)
                                    values.append(str(cell_value) if cell_value is not None else "")
                                    
                                    offset = offset + 1
                                target_value = self.DEFAULT_SPLITTER.join(values) if values else ""
                            else:
                                target_value = worksheet.cell(target_row, target_col).value if is_xlsx else worksheet.cell_value(target_row, target_col)
                            
                            dict_map[rule.field_name] = target_value
                        except Exception as ex:
                            raise ex
        return filename_no_ext, dict_map
    def _process_openpyxl(self, file_path: str, map: Dict[str, List[Rule]]) -> Tuple[str, dict]:
            """ 处理 .xlsx 文件 """
            src_wb = load_workbook(file_path)
            src_ws = src_wb.active
            Logger.debug(f"[xlsx] 工作表名称: {src_wb.sheetnames}")
            return self._process_common(file_path, src_ws, map, is_xlsx=True)

    def _process_xlrd(self, file_path: str, map: Dict[str, List[Rule]]) -> Tuple[str, dict]:
        """ 处理 .xls 文件 """
        src_wb = xlrd.open_workbook(file_path)
        src_ws = src_wb.sheet_by_index(0)  # 默认取第一个工作表
        Logger.debug(f"[xls] 工作表名称: {src_wb.sheet_names()}")
        return self._process_common(file_path, src_ws, map, is_xlsx=False)

    @deprecated(reason="更新至process_single_excel使用新的pandas进行更灵活的文件处理（支持xls)", version="1.2.0")
    def process_single_excel_v1(self, file_path: str, map: Dict[str, List[Rule]]) -> Tuple[str, dict]:
        """
        支持处理 .xls 和 .xlsx 文件的通用函数
        返回结构：(文件名, {"字段名": "字段值"})
        """
        # 根据扩展名选择读取方式
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.xlsx':
            return self._process_openpyxl(file_path, map)
        elif ext == '.xls':
            return self._process_xlrd(file_path, map)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
    
    def parse_excel_files(self, files, map , file_type = None, progress = None, key_field_name = None):
        """
        解析excel文件夹，返回结构：
        {
        "key_field_value": {文件A k-v字典},
        "key_field_value": {文件B k-v字典},
        ...
        key_field_name：区分文件的key，默认为"filename"(目前不支持)
        目前仅支持按照映射，转换成des_field_name的字典，key_field_value为文件名
        }
        """
        global_po_mapping = {}
        Logger.info(f"📩 发现 {len(files)} 封{file_type}待处理文件")
        with GlobalThreadPool.get_executor() as executor:
            futures = [
                executor.submit(self.process_single_excel, filename, map)
                for filename in files
            ]
            
        for future in concurrent.futures.as_completed(futures):
            key_field_v, fields = future.result()
            if key_field_v:
                global_po_mapping[key_field_v] = fields
                Logger.debug(f"✅ {key_field_v}：解析结果：{global_po_mapping[key_field_v]}")

        return global_po_mapping

    def process_single_excel(self, file_path: str, map: Dict[str, List[Rule]]) -> Tuple[str, dict]:
        """处理Excel文件，返回文件名和处理后的数据"""
        filename_no_ext = os.path.splitext(os.path.basename(file_path))[0]
        
        # 读取Excel文件（获取所有工作表名称）
        with pd.ExcelFile(file_path) as excel:
            sheet_names = excel.sheet_names  # 获取所有工作表名称
            Logger.debug(f"[Excel] 工作表名称: {sheet_names}")
            
            # 读取第一个工作表
            df = pd.read_excel(excel, sheet_name=sheet_names[0])
        
        dict_map = {}
        rows, columns = df.shape
        
        # 遍历DataFrame的所有单元格
        for row_idx in range(rows):
            for col_idx, col_name in enumerate(df.columns):
                cell_value = df.iloc[row_idx, col_idx]
                
                if pd.notna(cell_value) and str(cell_value) in map:
                    for rule in map[str(cell_value)]:
                        values = []
                        offset = 0
                        try:
                            # 确定目标位置和偏移方向
                            if rule.dir == "row":
                                target_row = row_idx + 1
                                target_col = col_idx
                                max_limit = rows
                            elif rule.dir == "column":
                                target_row = row_idx
                                target_col = col_idx + 1
                                max_limit = columns
                            else:
                                continue  # 无效方向，跳过
                                
                            # readingMode优先级较高
                            if hasattr(rule, 'readingMode') and rule.readingMode == "readUntilBlank":
                                while True:
                                    pos_row = target_row + (offset if rule.dir == "row" else 0)
                                    pos_col = target_col + (offset if rule.dir == "column" else 0)
                                    
                                    # 检查边界
                                    if (rule.dir == "row" and pos_row >= rows) or \
                                    (rule.dir == "column" and pos_col >= len(df.columns)):
                                        break
                                    
                                    try:
                                        cell_value = df.iloc[pos_row, pos_col]
                                    except IndexError:
                                        break
                                    
                                    if pd.isna(cell_value) or cell_value == "":
                                        break
                                    
                                    values.append(str(cell_value))
                                    offset += 1
                                
                                target_value = self.DEFAULT_SPLITTER.join(values) if values else ""
                            
                            # 处理固定数量的单元格
                            elif hasattr(rule, 'count') and rule.count:
                                for i in range(rule.count):
                                    pos_row = target_row + (i if rule.dir == "row" else 0)
                                    pos_col = target_col + (i if rule.dir == "column" else 0)
                                    
                                    if (rule.dir == "row" and pos_row >= rows) or \
                                    (rule.dir == "column" and pos_col >= len(df.columns)):
                                        raise IndexError("偏移超出工作表范围")
                                    
                                    try:
                                        cell_value = df.iloc[pos_row, pos_col]
                                    except IndexError:
                                        cell_value = None
                                    
                                    values.append(str(cell_value) if pd.notna(cell_value) else "")
                                
                                target_value = self.DEFAULT_SPLITTER.join(values) if values else ""
                            
                            # 默认情况：读取单个单元格
                            else:
                                try:
                                    target_value = df.iloc[target_row, target_col]
                                except IndexError:
                                    target_value = None
                            
                            # 存储结果
                            if pd.notna(target_value):
                                dict_map[rule.field_name] = str(target_value)
                            else:
                                dict_map[rule.field_name] = ""
                        
                        except Exception as ex:
                            Logger.error(f"处理单元格({row_idx}, {col_idx})时出错: {str(ex)}")
                            continue
        
        return filename_no_ext, dict_map
    
    @staticmethod
    def get_checked_excel_sheets(file_paths: List[str], preset_sheets: List[str], default_fallback: str = "Sheet1"):
        """
        增强版Excel文件校验方法
        
        Args:
            file_paths: 需要检查的Excel文件路径列表
            preset_sheets: 必须存在的预设工作表列表
            default_fallback: 数据为空时的默认回退表名
        
        Returns:
            字典结构: {文件路径: {工作表名: 工作表对象}}
        
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 格式错误/表缺失/数据空且无法回退
        """
        sheet_maps = {}

        for file_path in file_paths:
            # 基础校验
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ('.xlsx', '.xlsm', '.xls'):
                raise ValueError(f"不支持的文件格式: {file_ext}")

            wb = None
            xlrd_book = None
            try:
                # 加载工作簿
                if file_ext == '.xls':
                    # 使用xlrd处理旧格式
                    xlrd_book = xlrd.open_workbook(file_path)
                    sheet_names = xlrd_book.sheet_names()
                    get_sheet = lambda name: xlrd_book.sheet_by_name(name)
                else:
                    # 使用openpyxl处理新格式
                    wb = load_workbook(file_path, read_only=True)
                    sheet_names = wb.sheetnames
                    get_sheet = lambda name: wb[name]

                # 检查必需表存在
                missing_sheets = set(preset_sheets) - set(sheet_names)
                if missing_sheets:
                    raise ValueError(f"文件 {file_path} 缺失必需表: {', '.join(missing_sheets)}")

                # 构建有效表映射
                sheet_map = {}
                for req_sheet in preset_sheets:
                    try:
                        sheet = get_sheet(req_sheet)
                    except (KeyError, xlrd.biffh.XLRDError):
                        sheet = None

                    # 检查数据有效性
                    has_data = False
                    if sheet:
                        # 不同格式的数据检查逻辑
                        if file_ext == '.xls':
                            has_data = sheet.nrows > 0 and any(
                                sheet.cell_value(row, col) not in (None, "")
                                for row in range(sheet.nrows)
                                for col in range(sheet.ncols)
                            )
                        else:
                            has_data = any(
                                cell.value not in (None, "")
                                for row in sheet.iter_rows(min_row=1, max_row=1)
                                for cell in row
                            )

                    # 处理空数据回退
                    if not has_data:
                        Logger.warning(f"[{file_path}] {req_sheet} 数据为空，尝试回退到 {default_fallback}")
                        
                        try:
                            fallback_sheet = get_sheet(default_fallback)
                        except (KeyError, xlrd.biffh.XLRDError):
                            raise ValueError(f"回退表 {default_fallback} 不存在")

                        # 验证回退表数据
                        if file_ext == '.xls':
                            has_fallback_data = fallback_sheet.nrows > 0
                        else:
                            has_fallback_data = any(
                                cell.value for row in fallback_sheet.iter_rows()
                                for cell in row
                            )
                        
                        if not has_fallback_data:
                            raise ValueError(f"回退表 {default_fallback} 数据为空")
                        
                        sheet_map[req_sheet] = fallback_sheet
                    else:
                        sheet_map[req_sheet] = sheet

                sheet_maps[file_path] = sheet_map

            except Exception as e:
                # 资源清理
                if wb: wb.close()
                if xlrd_book: xlrd_book.release_resources()
                raise RuntimeError(f"文件 {file_path} 校验失败: {str(e)}") from e

        return sheet_maps
    @staticmethod
    def get_excel_sheets(file_paths, preset_sheets=None, read_only=True, verbose=False):
        """
        获取Excel文件中所有工作表的句柄
        
        :param file_paths: 要处理的Excel文件路径列表
        :param preset_sheets: 预设工作表名称列表（用于校验警告）
        :param read_only: 是否使用只读模式优化大文件加载性能
        :param verbose: 是否显示详细加载日志
        :return: 嵌套字典结构 {文件绝对路径: {工作表名称: 工作表对象}}
        """
        sheet_maps = {}
        preset_sheets = preset_sheets or []

        for file_path in file_paths:
            abs_path = str(Path(file_path).absolute())
            if verbose:
                Logger.info(f"⏳ 开始加载文件: {abs_path}")

            try:
                # 加载工作簿（自动关闭文件句柄）
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")  # 忽略openpyxl的警告
                    wb = load_workbook(
                        filename=abs_path,
                        read_only=read_only,
                        data_only=True,
                        keep_links=False  # 提高加载速度
                    )

                # 收集所有工作表句柄
                sheets = {}
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheets[sheet.title] = sheet
                    
                    # 记录非预设sheet警告
                    if preset_sheets and sheet.title not in preset_sheets:
                        msg = f"⚠️ 检测到非常规工作表 [{sheet.title}] 在文件 {Path(abs_path).name}"
                        Logger.debug(msg)

                # 保留工作簿引用避免被GC
                sheets["_workbook"] = wb  
                sheet_maps[abs_path] = sheets
                
                if verbose:
                    Logger.info(f"✅ 成功加载 {len(sheets)-1} 个工作表")

            except BadZipFile as e:
                error_msg = f"❌ 文件损坏无法打开: {Path(abs_path).name} ({str(e)})"
                Logger.error(error_msg)
            except Exception as e:
                error_msg = f"❌ 加载失败: {Path(abs_path).name}\n{traceback.format_exc()}"
                Logger.error(error_msg)

        return sheet_maps
    @staticmethod
    def load_excel_to_K_V(input_file, key_fields, progress = None):
        """
        将snt当前excel文件中active_sheet中关键字段keys——用于联系数据，的行写入内存{key_tuple,row}
        """
        try:
            output_ws = load_workbook(input_file).active

            snt_data = {}
            snt_gen = ExcelProcessor.excel_row_generator(
                output_ws,
                input_file,
                progress,
                key_fields,
                strict_flag=False
            )
            for row in snt_gen:
                key = tuple(row[field] for field in key_fields)
                if key in snt_data:
                    Logger.info(f"⚠️ 发现重复基准数据: {key}")
                snt_data[key] = row

            return snt_data
        except Exception as e:
            raise RuntimeError (f"❌ 内存加载{input_file}基准数据失败: {str(e)}")
    @staticmethod
    def read_excel_row(template_file_name, sheet_name, index, max_retries=3, retry_delay=1):
        """
        读取Excel文件中指定工作表的数据行，并返回表头与数据的键值对
        
        参数:
        template_file_name (str): Excel文件名(.xlsx或.xls)
        sheet_name (str): 工作表名称
        index (int): 要读取的数据行索引(0-based)
        max_retries (int): 最大重试次数(默认3)
        retry_delay (int): 重试延迟时间(秒，默认1)
        
        返回:
        dict: 包含表头和数据键值对的字典，失败时返回None
        """
        if not os.path.exists(template_file_name):
            raise FileNotFoundError(f"文件不存在: {template_file_name}")
        
        for attempt in range(max_retries + 1):
            try:
                # 读取Excel文件
                df = pd.read_excel(
                    template_file_name,
                    sheet_name=sheet_name,
                    header=0,  # 使用第一行作为列名
                    dtype=str,   # 将所有数据读取为字符串保持原始格式
                    keep_default_na=False  # 禁用默认NaN转换
                )
                
                # 验证行索引有效性
                if index < 0 or index >= len(df):
                    raise IndexError(f"行索引{index}超出范围(0-{len(df)-1})")
                
                # 获取表头和行数据
                headers = df.columns.tolist()
                row_values = df.iloc[index].tolist()
                
                # 创建键值对字典
                return {str(header): str(value) for header, value in zip(headers, row_values)}
            
            except Exception as e:
                Logger.debug(f"⚠️ 读取失败(尝试 {attempt+1}/{max_retries}): {str(e)}")
                if attempt < max_retries:
                    Logger.info(f"⏳ 等待 {retry_delay} 秒后重试...")
                    time.sleep(retry_delay)
                else:
                    Logger.debug(f"❌ 读取Excel文件中指定工作表的数据行失败: {str(e)}")
                    return None