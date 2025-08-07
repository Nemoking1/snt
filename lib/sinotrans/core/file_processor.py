from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook, Workbook
from sinotrans.utils.logger import Logger
from sinotrans.core.rule import Rule
from deprecated import deprecated
from typing import get_type_hints
from pathlib import Path
import pandas as pd
import openpyxl
import time
import os

class FileProcessor:
    """
    文件处理类,解析映射规则，创建输出文件,返回：
    {
    "src_field_name1":[rule1, rule2, ...],
    "src_field_name2":[rule1, rule2, ...],
    ...
    }
    """
    @staticmethod
    def read_files(folder_path, suffixes):
        """文件夹下对应后缀的文件绝对路径列表"""
        xls_files = []
        for to_f in os.listdir(folder_path): 
            if to_f.startswith("~$"):
                continue
            file_suffix = os.path.splitext(to_f)[1].lower()
            if file_suffix in suffixes:
                xls_file = os.path.join(folder_path, to_f)
                xls_files.append(xls_file)

        return xls_files
    @staticmethod
    def ensure_directories_exist(directories):
        """
        确保所有必要的目录存在，如果不存在则创建它们
        输入：directory_list  - 一个包含目录路径的"列表"
        """
        try:
            for directory in directories:
                if not os.path.exists(directory):
                    os.makedirs(directory)
        except Exception as e:
            raise Exception("❌ 创建目录失败")

    def create_newfile_by_template(template_file, target_file, additional_columns=None):
        """
        使用 pandas 复制模板中所有 sheet 的表头，并支持添加新列
        返回：absolute_path
        """
        try:
            # 读取所有 sheet 的第一行（自动识别 .xls/.xlsx）
            file = pd.ExcelFile(template_file)
            all_dfs = {}

            for sheet_name in file.sheet_names:
                df = pd.read_excel(file, sheet_name=sheet_name, nrows=0)# 读取第一个数据行，即可能读第二行pd.read_excel(file, sheet_name=sheet_name, nrows=1)
                if additional_columns:
                    for col in additional_columns:
                        df[col] = None
                all_dfs[sheet_name] = df

            # 确保保存为 .xlsx 格式
            if not target_file.lower().endswith('.xlsx'):
                target_file += '.xlsx'

            # 使用 ExcelWriter 写入多个 sheet
            with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
                for sheet_name, df in all_dfs.items():
                    df.to_excel(writer, index=False, sheet_name=sheet_name)

            return target_file

        except Exception as e:
            raise RuntimeError(f"❌ 处理失败: {str(e)}")
    @staticmethod
    def create_newfile_by_template_retryable(template_file, target_file, additional_columns=None, max_retries=3, retry_interval=5, start_index=0):
        """
        使用 pandas 复制模板中所有 sheet 的表头，并支持添加新列，支持自定义重试次数和重试间隔，默认重试3次
        返回：absolute_path
        """
        for attempt in range(1, max_retries + 1):
            try:
                # 读取所有 sheet 的第一行（自动识别 .xls/.xlsx）
                file = pd.ExcelFile(template_file)
            except Exception as e:
                Logger.debug(f"⚠️ 读取{template_file}文件失败 (尝试 {attempt}/{max_retries}): {e}")
                if attempt < max_retries:
                    Logger.debug(f"⏳ 等待 {retry_interval} 秒后重试...")
                    time.sleep(retry_interval)
                else:
                    raise RuntimeError(f"❌ 无法访问文件: {template_file}") from e
            try:
                all_dfs = {}
                for sheet_name in file.sheet_names:
                    df = pd.read_excel(file, sheet_name=sheet_name, nrows=start_index)# nrows=1读取第一个数据行，即可能读第二行pd.read_excel(file, sheet_name=sheet_name, nrows=1)
                    if additional_columns:
                        for col in additional_columns:
                            df[col] = None
                    all_dfs[sheet_name] = df

                # 确保保存为 .xlsx 格式
                if not target_file.lower().endswith('.xlsx'):
                    target_file += '.xlsx'

                # 使用 ExcelWriter 写入多个 sheet
                with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
                    for sheet_name, df in all_dfs.items():
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                return target_file
            except Exception as e:
                raise RuntimeError(f"❌ 处理失败: {str(e)}")
    @staticmethod
    def apply_default_style(output_ws):
        """初始化Excel样式"""
        wb = output_ws.parent  # 获取工作簿对象
        
        # 修复：直接使用 wb.style_names 获取样式名称列表
        if "header_style" not in wb.style_names:
            # 样式不存在则创建
            header_style = openpyxl.styles.NamedStyle(name="header_style")
            header_style.font = openpyxl.styles.Font(name="Calibri", bold=True, color="FFFFFF")
            header_style.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="4F81BD")
            wb.add_named_style(header_style)  # 添加到工作簿
        
        # 应用冻结窗格和表头样式
        output_ws.freeze_panes = "A2"
        for cell in output_ws[1]:
            cell.style = "header_style"  # 使用样式名称而不是对象
        
        # 设置固定列宽（优化版）
        # 使用更高效的方式设置所有列的宽度
        for col_letter in output_ws.column_dimensions:
            output_ws.column_dimensions[col_letter].width = 40
        
        # 隔行填充背景色（优化版）
        even_fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='C8D7E9', end_color='C8D7E9')
        base_font = openpyxl.styles.Font(name='Calibri', size=11)
        
        # 只遍历有数据的行
        for row in output_ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = base_font
                if cell.row % 2 == 0:
                    cell.fill = even_fill
    @staticmethod
    def save_file_retryable(file, data=None, is_format_applied=False, save_mode='auto', is_append=False, sheet_name=None, max_retries=5, retry_interval=5, output_wb=None):
        """
        通用的文件保存函数，支持重试机制和多种数据格式
        
        参数:
        file: 文件绝对路径
        data: 要保存的数据（二进制数据或有序的行数据列表）
        format_wb,format_ws: 带格式的wb/s (适用于excel，如果有，优先使用)
        save_mode: 保存模式 ('binary', 'excel_xls', 'excel_xlsx', 'auto')
        sheet_name: Excel 工作表名称（仅当保存行数据时使用）
        max_retries: 最大重试次数
        retry_interval: 重试间隔时间（秒）
        """
        for attempt in range(1, max_retries + 1):
            try:
                if sheet_name is None and output_wb is not None:
                    sheet_name = output_wb.active.title
                # 预保存工作簿
                if output_wb:
                    output_wb.save(file)
                    Logger.debug(f"✅ xlsx文件已预保存")
                # 自动检测保存模式
                if save_mode == 'auto':
                    if file.lower().endswith('.xlsx'):
                        save_mode = 'excel_xlsx'
                    elif file.lower().endswith('.xls'):
                        save_mode = 'excel_xls'
                    else:
                        save_mode = 'binary'
                # 执行保存操作
                if save_mode == 'binary':
                    if data is None:
                        raise ValueError("❌ 二进制保存模式需要提供data参数")
                    
                    if not isinstance(data, (bytes, bytearray)):
                        raise TypeError("❌ 二进制模式需要字节类型数据")
                    
                    # 二进制写入, 如果 original_eml_file 不存在，open(..., 'wb') 会自动创建该文件。
                    with open(file, 'wb') as f:
                        f.write(data)
                    Logger.info(f"✅ 二进制数据已保存到: {file}")
                            
                elif save_mode in ['excel_xls', 'excel_xlsx']:
                    # 如果有行数据，则用 pandas 统一处理
                    if data and isinstance(data, list):
                        df = pd.DataFrame(data)

                        # 确保目标路径是 .xlsx
                        new_file = file if file.lower().endswith('.xlsx') else file + '.xlsx'

                        # 追加模式判断
                        if is_append:
                            # 严格检查：文件必须存在且为xlsx格式
                            if not os.path.exists(new_file):
                                raise FileNotFoundError(f"❌ 追加模式错误：目标文件 '{new_file}' 不存在")
                            
                            if not new_file.lower().endswith('.xlsx'):
                                raise ValueError(f"❌ 追加模式错误：文件 '{new_file}' 不是.xlsx格式")
                            
                            try:
                                # 读取现有文件内容
                                existing_df = pd.read_excel(new_file, sheet_name=sheet_name, header=None)
                                
                                # 合并新旧数据
                                combined_df = pd.concat([existing_df, df], ignore_index=True)
                                
                                # 使用新数据覆盖写入指定sheet
                                with pd.ExcelWriter(new_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                                    
                            except Exception as e:
                                Logger.error(f"❌ 追加数据失败: {str(e)}")
                                raise
                        else:
                            # 使用 pandas 写入 [.xlsx](file://d:\project\EmlAidata\template.xlsx) 文件
                            with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                        Logger.info(f"✅ 成功写入 {len(data)} 行数据到 '{sheet_name}' 工作表")
                    
                    # 情况2：data是二进制数据 → 直接写入文件
                    elif isinstance(data, bytes):
                        with open(file, 'wb') as f:
                            f.write(data)
                            
                    # 情况2：因为数据为None, 只创建新文件，若文件已存在，则不处理
                    else:
                        if not os.path.exists(file):
                            wb = Workbook()
                            wb.save(file)
                            Logger.info(f"✅ 创建新的Excel文件: {file}")
                        else:
                            Logger.debug(f"❌ 文件已存在: {file}，数据：{data} 不写入")
                        # else:
                        #     df = pd.read_excel(file)  # 自动识别 .xls/.xlsx
                        #     df.to_excel(file + '.xlsx', index=False)  # 转换为 .xlsx
                        #     os.remove(file)  # 删除原始文件
                        #     Logger.info(f"✅ 将文件 {file} 转换为 .xlsx 格式")

                if is_format_applied:
                    wb, ws = FileProcessor.load_wordbook_retryable(file, sheet_name=sheet_name)
                    FileProcessor.apply_default_style(ws)
                    wb.save(file)
                    Logger.debug(f"✅ xlsx文件格式已保存")

                return True
                
            except PermissionError as e:
                Logger.error(f"⚠️ 文件被占用 (尝试 {attempt}/{max_retries}): {e}")
                if attempt < max_retries:
                    Logger.info(f"⏳ 等待 {retry_interval} 秒后重试...")
                    time.sleep(retry_interval)
                else:
                    raise RuntimeError(f"❌ 无法访问文件: {file}") from e
                    
            except InvalidFileException as e:
                # 对于无效Excel文件不重试
                raise RuntimeError(f"❌ 无效的Excel文件格式: {file}") from e
                
            except Exception as e:
                Logger.error(f"⚠️ 保存失败 (尝试 {attempt}/{max_retries}): {type(e).__name__} - {e}")
                if attempt < max_retries:
                    Logger.info(f"⏳ 等待 {retry_interval} 秒后重试...")
                    time.sleep(retry_interval)
                else:
                    raise RuntimeError(f"❌ 文件保存失败: {file}") from e

    @deprecated(reason="更新至create_newfile_by_template使用新的pandas进行更灵活的文件处理（支持xls)", version="1.2.0")
    def create_newfile_by_template_v1(template_file_name, target_file_name, column_names = None):
            """
            创建新文件，并复制模板表头
            可选：添加新列
            """
            # 加载模板并获取表头（第一行数据）
            header_row = next(load_workbook(template_file_name).active.iter_rows(max_row=1, values_only=True))  # 提取第一行数据
            # 添加新列
            header_row = list(header_row)
            if not column_names:
                header_row.extend(column_names)
            # 创建新工作簿并写入表头
            new_wb = Workbook()
            new_sheet = new_wb.active
            new_sheet.append(header_row)
            
            new_wb.save(target_file_name)
            return new_wb
    @staticmethod
    def parse_rule(value, prefix_separator, condition_separator, key_value_separator):
        """
        解析"指定单行"中"指定格式分隔符"的键值对格式的字符串，将其解析成：
        Rule对象
        """
        # 初始化Rule对象
        rule = Rule(field_name=value.split('|')[0].strip())
        # 获取Rule对象的属性类型
        type_hints = get_type_hints(type(rule))
        # 解析参数键值对
        if prefix_separator in value:
            # 先取条件字符串
            params_str = value.split(prefix_separator, 1)[1]
            # 遍历条件
            for param in params_str.split(condition_separator):
                # param不要去掉末尾空格，不然分隔符为空格的时候就会被误删！
                param = param
                if key_value_separator not in param:
                    continue
                # 分割当前条件，提取键值对
                k, v = param.split(key_value_separator, 1)
                k = k.strip()
                # 判断属性是否存在，存在则尝试转换类型并赋值
                if hasattr(rule, k):
                    attr_type = type_hints[k]
                    try:
                        converted_v = attr_type(v)
                        setattr(rule, k, converted_v)
                    except (TypeError, ValueError) as e:
                        raise ValueError(f"❌ 参数 {k} 的值 {v} 无法转换为正确类型: {e}")
        return rule
    @staticmethod
    def parse_conf(file_name, splitter):
        """
        解析"指定文件"中"指定格式分隔符"的键值对格式的所有行，将其解析成：
        {name1,name2,name3}
        """
        with open(file_name, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or splitter not in line:
                    return line
                values = line.split(splitter)
        return values
    @staticmethod
    def parse_mapping_dict(file_name, splitter, prefix_separator, condition_separator, key_value_separator):
        """
        解析"指定文件"中"指定格式分隔符"的键值对格式的所有行，将其解析成：
        {
        "src_field_name1":rule1,
        "src_field_name2":rule2,
        ...
        }
        """
        mapping = {}
        with open(file_name, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or splitter not in line:
                    continue
                key, value = line.split(splitter, 1)
                rule = FileProcessor.parse_rule(value, prefix_separator, condition_separator, key_value_separator)
                mapping[key.strip()] = rule
        return mapping
    @staticmethod
    def parse_mapping_dict_of_list(file_name, splitter, prefix_separator, condition_separator, key_value_separator):
        """
        解析"指定文件"中"指定格式分隔符"的键值对格式的所有行，将其解析成：
        {
        "src_field_name1":[rule1, rule2, ...],
        "src_field_name2":[rule1, rule2, ...],
        ...
        }
        """
        mapping = {}
        with open(file_name, 'r', encoding='utf-8') as f:
            for line in f:
                # 去除空白字符
                line = line.strip()
                # 根据splitter分隔符判断该行是否有效——键值对格式，否则，跳过
                if not line or splitter not in line:
                    continue
                # 根据splitter分隔符，获取键值对
                key, value = line.split(splitter, 1)
                # 映射对象存在，则追加rule, 否则，新建
                if key not in mapping:
                    mapping[key] = []
                rule = FileProcessor.parse_rule(value, prefix_separator, condition_separator, key_value_separator)
                mapping[key.strip()].append(rule)
        return mapping
    @staticmethod
    def file_generator(file_path, clp_file_content_map, to_file_content_map, progress, required_keys=None):
        """遍历文件内容映射表，生成有效文件绝对路径列表的生成器，并检查文件名是否存在"""
        Logger.debug("📂 开始遍历文件内容映射表")
        
        # 获取所有唯一文件名（合并两个映射表的key）
        all_files = set(clp_file_content_map.keys()).union(set(to_file_content_map.keys()))
        
        progress.init_main_progress(len(all_files))
        for file_name in all_files:
            # 更新进度
            progress.update()
            
            try:
                # 检查文件是否在两个映射表中都存在
                in_clp = file_name in clp_file_content_map
                in_to = file_name in to_file_content_map
                
                if not in_clp or not in_to:
                    Logger.debug(f"- 跳过无效文件：{file_name}（舱单文件或clp文件缺失）")
                    continue
                    
                # 构建文件数据字典
                file_data = {
                    "file_name": file_name,
                    "in_clp": in_clp,
                    "in_to": in_to,
                    "clp_content": clp_file_content_map.get(file_name),
                    "to_content": to_file_content_map.get(file_name)
                }
                
                # 必填字段检查
                if required_keys:
                    missing_keys = [
                        key for key in required_keys 
                        if file_data["clp_content"].get(key) in (None, "") \
                        or file_data["to_content"].get(key) in (None, "")
                    ]
                    if missing_keys:
                        Logger.debug(f"跳过无效文件：{file_name}，缺失字段：{', '.join(missing_keys)}")
                        continue
                
                # # 数据清洗：字符串去空格
                # for key, value in file_data.items():
                #     if isinstance(value, str):
                #         file_data[key] = value.strip()
                
                yield os.path.join(file_path, file_name)
                
            except Exception as e:
                Logger.error(f"❌ 文件 {file_name} 处理失败: {str(e)}")
                continue
    @staticmethod
    def write_rows_to_files(add_rows):
        """
        将 add_rows 中的数据写入对应的 Excel 文件中。
        
        参数:
            add_rows (dict): {文件名: 需要添加的行列表}
        """
        for file_path, rows in add_rows.items():
            try:
                if os.path.exists(file_path):
                    # 如果文件存在，加载现有工作簿
                    wb = load_workbook(file_path)
                else:
                    raise FileNotFoundError(f"文件 {file_path} 不存在")

                ws = wb.active
                # 写入数据行
                for row in rows:
                    ws.append(list(row.values()))
                # 保存文件
                wb.save(file_path)
                Logger.info(f"✅ 数据已成功写入文件: {file_path}")
            except Exception as e:
                Logger.error(f"❌ 写入文件 {file_path} 时出错: {str(e)}")
    @staticmethod
    def load_wordbook_retryable(file, sheet_name=None, max_retries=3, retry_interval=5):
        for attempt in range(1, max_retries + 1):
            try:
                output_wb = load_workbook(file)
                if sheet_name is None:
                    output_ws = output_wb.active
                else:
                    output_ws = output_wb[sheet_name]
                return output_wb, output_ws
            except Exception as e:
                Logger.info(f"读取{file}文件失败 (尝试 {attempt}/{max_retries}): {e}")
                if attempt < max_retries:
                    Logger.info(f"等待 {retry_interval} 秒后重试...")
                    time.sleep(retry_interval)
                else:
                    raise RuntimeError(f"无法访问文件: {file}") from e
