from datetime import datetime
import os
from openpyxl import load_workbook
import re
import sys
import concurrent.futures
from sinotrans.core import FileParser,ExcelProcessor,EmlParser
from sinotrans.utils import Logger,GlobalThreadPool, ProgressManager

# # 配置路径参数：os.path.abspath(__file__)
# # 打包替换启动路径为：root_path3 = os.path.dirname(os.path.realpath(sys.executable))
# timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
# CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))   
# TARGET_PATH = os.path.join(CURRENT_DIR, "target")
# CONFIG_PATH = os.path.join(CURRENT_DIR, "conf")
# DEBUG_PATH = os.path.join(CURRENT_DIR, "logs")
# # 文件路径配置
# RESOURCE_FILE = os.path.join(CURRENT_DIR, "input_data.xlsx")
# TEMPLATE_FILE = os.path.join(CURRENT_DIR, "template.xlsx")
# TARGET_FILE = os.path.join(TARGET_PATH, f"output_result_{timestamp}.xlsx")
# MAPPING_FILE = os.path.join(CONFIG_PATH, "mapping.txt")
# FIXED_MAPPING_FILE = os.path.join(CONFIG_PATH, "fixed_mapping.txt")
# EMAIL_MAPPING_FILE = os.path.join(CONFIG_PATH, "email_mapping.txt")
# CONTAINER_TYPES = ["20GP", "40GP", "40HQ"]
# PO_NAME="PO号"
# CONTAINER_TYPE_NAME="集装箱类型"
# 配置路径参数：os.path.abspath(__file__)
# 打包替换启动路径为：root_path3 = os.path.dirname(os.path.realpath(sys.executable))
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))  
TARGET_PATH = os.path.join(CURRENT_DIR, "target")
CONFIG_PATH = os.path.join(CURRENT_DIR, "conf")
RESPONSE_PATH = os.path.join(CURRENT_DIR, "res")
SNT_PATH = os.path.join(CURRENT_DIR, "snt")
REPORT_PATH = os.path.join(CURRENT_DIR, "report")
DEBUG_PATH = os.path.join(CURRENT_DIR, "logs")

SHEET_CONFIG_FILE = os.path.join(CONFIG_PATH, "sheet_config.txt")

FIXED_MAPPING_FILE = os.path.join(CONFIG_PATH, "fixed_mapping.txt")
BC4_REPORT_MAPPING_FILE = os.path.join(CONFIG_PATH, "bc4_report_mapping.txt")
PENDING_PO_MAPPING_FILE = os.path.join(CONFIG_PATH, "pending_po_mapping.txt")
RESPONSE_MAPPING_FILE = os.path.join(CONFIG_PATH, "response_mapping.txt")

TEMPLATE_FILE = os.path.join(CURRENT_DIR, "template.xlsx")
TARGET_FILE = os.path.join(TARGET_PATH, f"output_result_{timestamp}.xlsx")

REQUIRED_FIELDS=[
    "folder",
    "po",
    "lot",
]
Logger(debug_path=DEBUG_PATH)

FileParser.ensure_directories_exist(directories = [
    TARGET_PATH,
    CONFIG_PATH,
    SNT_PATH,
    RESPONSE_PATH,
])
GlobalThreadPool.initialize(
    max_workers=16,
    thread_name_prefix='AutoTOThreadPool',
    initializer=lambda: Logger.debug("AutoTOThreadPool initialized"),
    initargs=()
)
# 预定义配置
DEFAULT_FALLBACK_SHEET = "Sheet1"  # 默认回退工作表

def validate_and_get_sheets(wb, file_path, sheet_names):
    """验证工作簿并返回有效工作表映射"""
    # 检查必需工作表是否存在
    existing_sheets = set(wb.sheetnames)
    missing_sheets = set(sheet_names) - existing_sheets
    if missing_sheets:
        raise ValueError(f"文件 {file_path} 缺失必需工作表: {', '.join(missing_sheets)}")

    # 获取有效工作表（数据不为空时优先，否则使用回退表）
    sheet_map = {}
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        # 检查数据是否为空（至少有一个非空单元格）
        has_data = any(
            cell.value not in (None, "")
            for row in ws.iter_rows(min_row=1, max_row=1)  # 检查首行
            for cell in row
        )
        
        if not has_data:
            Logger.warning(f"工作表 {sheet_name} 数据为空，尝试回退到{DEFAULT_FALLBACK_SHEET}")
            if DEFAULT_FALLBACK_SHEET in wb.sheetnames:
                fallback_ws = wb[DEFAULT_FALLBACK_SHEET]
                if any(cell.value for cell in next(fallback_ws.iter_rows())):
                    sheet_map[sheet_name] = fallback_ws
                else:
                    raise ValueError(f"回退表{DEFAULT_FALLBACK_SHEET}也为空")
            else:
                raise ValueError(f"无有效数据且找不到回退表{DEFAULT_FALLBACK_SHEET}")
        else:
            sheet_map[sheet_name] = ws
    
    return sheet_map
# def process_fields(row_data, global_po_mapping):
#     """复制源行数据，并结合邮件中集装箱类型及个数，生成多个新行"""
#     try:
#         po = str(row_data[PO_NAME])
#         if po not in global_po_mapping:
#             return []
#         rows_to_add = []
#         quantity = 1
#         container_str = global_po_mapping[po].pop(CONTAINER_TYPE_NAME, None)
#         if container_str is None:
#             raise ValueError(f"❌ 未找到与 {po} 相关联的集装箱类型！")
#         container_types = container_str.split(",")
#         for type_str in container_types:
#             match = re.search(r'(\d+)$', type_str)
#             if match:
#                 quantity = int(match.group(1))
#                 type_part = type_str[:match.start(1)]
#                 type = type_part[:-1].replace("HC", "HQ").replace("GC", "GP").replace("STD", "GP")
#             for _ in range(quantity):
#                 new_row = row_data.copy()
#                 new_row[CONTAINER_TYPE_NAME] = type
#                 rows_to_add.append(new_row)
#     except Exception as e:
#         print(f"❌ 根据原行数据，生成新行数据时发生错误: {str(e)}")
#         raise
#     return rows_to_add
def map_fields(report_rows, response_rows, snt_rows, fixed_mapping, snt_mapping, bc4_report_mapping, response_mapping):
    """根据模板列顺序对生成的新行数据进行排序"""
    mapped_rows = []
    for row in report_rows:
        mapped_row = {}
        mapped_row.update(ExcelProcessor.fixed_mapping(fixed_mapping))
        # mapped_row.update(ExcelProcessor.column_mapping(row, snt_mapping))
        mapped_row.update(ExcelProcessor.column_mapping(row, bc4_report_mapping))
        # mapped_row.update(ExcelProcessor.column_mapping(row, response_mapping))
        mapped_rows.append(mapped_row)
    for row in response_rows:
        mapped_row = {}
        mapped_row.update(ExcelProcessor.fixed_mapping(fixed_mapping))
        mapped_row.update(ExcelProcessor.column_mapping(row, response_mapping))
        mapped_rows.append(mapped_row)
    for row in snt_rows:
        mapped_row = {}
        mapped_row.update(ExcelProcessor.fixed_mapping(fixed_mapping))
        mapped_row.update(ExcelProcessor.column_mapping(row, snt_mapping))
        mapped_rows.append(mapped_row)
    return mapped_rows
def process_resource_row(report_rows, response_rows, snt_rows, ns_output, fixed_mapping, snt_mapping, bc4_report_mapping, response_mapping):
    try:
        """处理源行数据，生成新行数据并返回"""
        mapped_rows = map_fields(report_rows, response_rows, snt_rows, fixed_mapping, snt_mapping, bc4_report_mapping, response_mapping)
        sorted_rows = ExcelProcessor.sort_generated_rows(mapped_rows, ns_output)
        return sorted_rows
    except Exception as e:
        Logger.error(f"❌ 处理源行数据失败: {str(e)}")
        raise
def process_resource_data(ns_output, snt_data_generator, report_data_generator, response_data_generator, progress, fixed_mapping, snt_mapping, bc4_report_mapping, response_mapping):
    new_rows = [] 
    try:
        Logger.info("📋 正在处理资源数据...")
        with GlobalThreadPool.get_executor() as executor:
            futures =[]
            # 同时遍历三个生成器，每次取一行传入 process_resource_row
            for snt_row, report_row, response_row in zip(snt_data_generator, report_data_generator, response_data_generator):
                future = executor.submit(
                    process_resource_row,
                    snt_row, report_row, response_row, ns_output,
                    fixed_mapping, snt_mapping, bc4_report_mapping, response_mapping
                )
                futures.append(future)
        progress.close()
        # 等待所有任务完成，设置超时时间
        done, not_done = concurrent.futures.wait(futures, timeout = 60)

        for future in futures:
            # 如果某些线程抛出异常，在调用 future.result() 时仍会触发异常。
            result = future.result()
            if result:
                new_rows.extend(result)
        Logger.info(f"✅ 扫描到{len(futures)}行非空数据，生成 {len(new_rows)} 行数据")
        return new_rows
    except Exception as e:
        Logger.error(f"❌ 处理资源数据失败: {str(e)}")
        raise

def main():
    try:
        # 读取映射关系：sheer_config,pending_po_map,response_map,bc4_report_map
        sheet_names = FileParser.parse_conf(SHEET_CONFIG_FILE,',')
        fixed_mapping = FileParser.parse_mapping_dict(FIXED_MAPPING_FILE,':', '|', ',', '=')   # 模板值映射
        snt_mapping = FileParser.parse_mapping_dict_of_list(PENDING_PO_MAPPING_FILE,':', '|', ',', '=')
        response_mapping = FileParser.parse_mapping_dict(RESPONSE_MAPPING_FILE,':', '|', ',', '=')
        bc4_report_mapping = FileParser.parse_mapping_dict(BC4_REPORT_MAPPING_FILE,':', '|', ',', '=')
        Logger.info("✅ 读取映射文件成功！")
    except Exception as e:
        Logger.error(f"❌ 映射文件读取失败: {str(e)}")
        return 
    try:
        # 获取snt、报文、业务表的文件名列表
        snt_files = FileParser.read_files(SNT_PATH, [".xlsx",".xls"])
        report_files = FileParser.read_files(REPORT_PATH, [".xlsx",".xls"])
        response_files = FileParser.read_files(RESPONSE_PATH, [".xlsx",".xls"])
        Logger.info("✅ 获取snt、报文、业务回复的文件成功！")
    except Exception as e:
        Logger.error(f"❌ 获取snt、报文、业务回复的文件失败: {str(e)}")
        return 
    
    # TODO 批量，需要文件名的设置
    # try:
    #     # 根据映射文件中的读取规则，提取装箱计划、舱单对应字段值
    #     # to_files = parse_excel_files(to_files, to_mapping)
    #     excel_processor = ExcelProcessor()
    #     snt_file_content_map = excel_processor.parse_excel_files(files=snt_files, map=clp_map, progress=None, key_field_name=None)
    #     report_file_content_map = excel_processor.parse_excel_files(files=report_files, map=to_map, progress=None, key_field_name=None)
    #     Logger.info("✅ 提取装箱计划、舱单对应字段值成功！")
    # except Exception as e:
    #     Logger.error(f"❌ 提取装箱计划、舱单对应字段值失败: {str(e)}")
    #     return 
    try:
        # 检查多个Excel文件是否包含指定的所有工作表且数据不为空
        ExcelProcessor.check_excel_sheets(snt_files + response_files, sheet_names)
        # 获取新工作簿和输出工作表
        nf_output = FileParser.create_newfile_by_template(TEMPLATE_FILE, TARGET_FILE, list("feedback"))
        Logger.info("✅ 创建模板文件")
        # 仅读取源数据的值/表头
        Logger.info("📋 开始读取源数据......")
        # 因为不是批量所以仅读取第一个文件，作为snt-
        input_workbooks = {
            "snt": load_workbook(snt_files[0], read_only=True),
            "report": load_workbook(report_files[0], read_only=True),
            "response": load_workbook(response_files[0], read_only=True)
        }
         # 建立工作表映射关系
        sheet_mapping = []
        for file_type, wb in input_workbooks.items():
            for sheet_name in wb.sheetnames:
                sheet_mapping.append((
                    wb[sheet_name],  # 输入工作表
                    nf_output[sheet_name],  # 输出工作表
                    file_type  # 文件类型标识
                ))

        Logger.info("✅ 所有输入文件验证通过，读取源数据成功！")
    except Exception as e:
        Logger.error(f"❌ 获取数据失败: {str(e)}")
        if os.path.exists(TARGET_FILE):
            os.remove(TARGET_FILE)
            Logger.error("❌ 获取数据失败，已删除目标文件。")
        return 
    try:
        progress = ProgressManager()
        for input_ws, output_ws, file_type in sheet_mapping:
            # 使用生成器逐行处理，采用生成器模式避免一次性加载所有数据到内存中
            data_generator = ExcelProcessor.excel_row_generator(
                input_ws,
                file_type,
                progress,
                required_columns=REQUIRED_FIELDS
            )
            new_rows = process_resource_data(output_ws, snt_data_generator, report_data_generator, response_data_generator, progress, fixed_mapping, snt_mapping, bc4_report_mapping, response_mapping)
            list(map(lambda row: output_ws.append(row), new_rows))
        progress.close()
        Logger.info("✅ 处理完成！")
    except Exception as e:
        Logger.error(f"❌ 源数据处理失败: {str(e)}")
        if os.path.exists(TARGET_FILE):
            os.remove(TARGET_FILE)
            Logger.debug("❌ 源数据处理失败，已删除目标文件。")
        return 
    try:
        # list(map(lambda row: ns_output.append(row), new_rows))
        nf_output.save(TARGET_FILE)
        Logger.info(f"🎉 结果已保存至: {TARGET_FILE}")
    except Exception as e:
        Logger.error(f"❌ 结果保存失败: {str(e)}")
        if os.path.exists(TARGET_FILE):
            os.remove(TARGET_FILE)
            Logger.debug("❌ 结果保存失败，已删除目标文件。")
        return

if __name__ == "__main__":
    main()
    input("Press Enter to exit...")