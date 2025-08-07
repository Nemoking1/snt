from datetime import datetime
import os
import sys
import concurrent.futures
import openpyxl
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from sinotrans.core import FileProcessor, ExcelProcessor
from sinotrans.utils import Logger, GlobalThreadPool, ExcelProgressTracker
import warnings
import traceback
import threading

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class AutoSntProcessor:
    # 预定义配置
    # 如果需要的表不存在，默认回退到默认表(可多个)
    DEFAULT_SHEET = "default_sheet"
    # 需要的表
    REQUIRED_SHEET = "required_sheet"
    # 用于关联不同文件中的行数据，表中必须存在的字段
    KEY_FIELDS = "key_fields"
    # 用于检验表中数据的有效性，通常和strice_flag配合使用
    REQUIRED_FIELDS = "required_fields"

    def __init__(self):
        # 初始化路径配置os.path.dirname(os.path.realpath(sys.executable))os.path.abspath(__file__)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.current_dir = os.path.dirname(os.path.abspath(__file__))
        self._init_paths()
        self._init_logger()
        self._init_thread_pool()
        self._init_styles()

    def _init_paths(self):
        """初始化所有路径配置"""
        self.target_path = os.path.join(self.current_dir, "target")
        self.config_path = os.path.join(self.current_dir, "conf")
        self.snt_path = os.path.join(self.current_dir, "snt")
        self.response_path = os.path.join(self.current_dir, "res")
        self.report_path = os.path.join(self.current_dir, "report")
        
        self.template_file = os.path.join(self.current_dir, "template.xlsx")
        self.target_file = os.path.join(self.target_path, f"PendingPoSnt_{self.timestamp}.xlsx")

        self.sheet_config_file = os.path.join(self.config_path, "sheet_config.txt")
        self.fixed_mapping_file = os.path.join(self.config_path, "fixed_mapping.txt")
        self.pending_po_mapping_file = os.path.join(self.config_path, "pending_po_mapping.txt")
        self.response_mapping_file = os.path.join(self.config_path, "response_mapping.txt")
        self.report_mapping_file = os.path.join(self.config_path, "report_mapping.txt")
        
        FileProcessor.ensure_directories_exist([
            self.target_path, self.config_path,
            self.snt_path, self.response_path, self.report_path
        ])

    def _init_logger(self):
        """初始化日志系统"""
        debug_path = os.path.join(self.current_dir, "logs")
        Logger(debug_path=debug_path)

    def _init_thread_pool(self):
        """初始化全局线程池"""
        GlobalThreadPool.initialize(
            max_workers=16,
            thread_name_prefix='AutoSNTThreadPool'
        )
    def _init_styles(self):
        """初始化Excel样式"""
        self.header_style = openpyxl.styles.NamedStyle(name="header_style")
        self.header_style.font = openpyxl.styles.Font(name="Calibri", bold=True, color="FFFFFF")
        self.header_style.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="4F81BD")
        # self.header_style.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    def _style_apply(self, output_ws):
        output_ws.freeze_panes = "A2"
        for cell in output_ws[1]:
            cell.style = self.header_style
        # 设置固定列宽（所有列宽度为15）
        list(map(lambda col: setattr(output_ws.column_dimensions[col], 'width', 20), ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']))
        # 隔行填充背景色
        fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='C8D7E9', end_color='C8D7E9')
        for row_num in range(2, output_ws.max_row + 1):  # 从第2行开始（第1行是表头）
            if row_num % 2 == 0:
                for col_num in range(1, output_ws.max_column + 1):
                    cell = output_ws.cell(row=row_num, column=col_num)
                    cell.fill = fill
                    cell.font = openpyxl.styles.Font(name='Calibri', size=11)
            else:
                for col_num in range(1, output_ws.max_column + 1):
                    cell = output_ws.cell(row=row_num, column=col_num)
                    cell.font = openpyxl.styles.Font(name='Calibri', size=11)


    def _thread_safe_process_sheet(self, sheet_name, template_wb):
        """线程安全的工作表处理方法"""
        try:
            # 创建临时工作簿副本
            thread_wb = openpyxl.Workbook()
            template_ws = template_wb[sheet_name]
            new_ws = thread_wb.create_sheet(sheet_name)
            
            # 复制表头
            for row in template_ws.iter_rows():
                new_ws.append([cell.value for cell in row])
            
            # 执行实际处理（操作临时工作簿）
            success = self._process_single_sheet(sheet_name, thread_wb)
            
            # 提取处理后的数据
            processed_data = []
            for row in new_ws.iter_rows(min_row=2):  # 跳过标题行
                processed_data.append([cell.value for cell in row])
                
            return (success, processed_data)
        except Exception as e:
            Logger.error(f"线程处理异常: {traceback.format_exc()}")
            return (False, None)
    def _load_mappings(self):
        """加载所有映射配置"""
        try:
            Logger.info("📋 开始处理映射文件......")
            sheet_conf = FileProcessor.parse_mapping_dict(self.sheet_config_file,':', '|', ',', '=')
            self.default_fallback_sheets = sheet_conf.get(self.DEFAULT_SHEET).field_name.split(",")
            self.key_fields = sheet_conf.get(self.KEY_FIELDS).field_name.split(",")
            self.required_fields = sheet_conf.get(self.REQUIRED_FIELDS).field_name.split(",")
            self.sheet_names = sheet_conf.get(self.REQUIRED_SHEET).field_name.split(",")

            self.fixed_mapping = FileProcessor.parse_mapping_dict(self.fixed_mapping_file,':', '|', ',', '=')   # 模板值映射
            self.snt_mapping = FileProcessor.parse_mapping_dict_of_list(self.pending_po_mapping_file,':', '|', ',', '=')
            self.response_mapping = FileProcessor.parse_mapping_dict_of_list(self.response_mapping_file,':', '|', ',', '=')
            self.report_mapping = FileProcessor.parse_mapping_dict_of_list(self.report_mapping_file,':', '|', ',', '=')

            Logger.info("✅ 映射文件加载成功")
        except Exception as e:
            Logger.error(f"❌ 映射文件加载失败: {str(e)}")
            raise

    def _validate_input_files(self):
        """验证输入文件完整性"""
        try:
            self.snt_files = FileProcessor.read_files(self.snt_path, [".xlsx", ".xls"])
            self.response_files = FileProcessor.read_files(self.response_path, [".xlsx", ".xls"])
            self.report_files = FileProcessor.read_files(self.report_path, [".xlsx", ".xls"])

            # 校验所有文件的工作表结构
            all_files = self.snt_files + self.response_files + self.report_files
            # 预设模板文件检查（保持严格校验）
            self.sheet_maps = ExcelProcessor.get_excel_sheets(
                file_paths=all_files,
                preset_sheets=self.sheet_names,  # 用于生成警告信息
                read_only=True,
                verbose=False
            )
            Logger.info("✅ 文件验证通过")
        except Exception as e:
            Logger.error(f"❌ 文件验证失败: {str(e)}")
            raise

    def _get_valid_sheet(self, file_sheets, sheet_name):
        """
        动态获取有效工作表
        参数：
        file_sheets: 工作表字典
        sheet_name: 目标工作表名
        返回值：
        - 工作表对象
        - 是否有使用回退表
        - 使用的回退表名
        """
        # 首选目标目标表名
        if sheet_name in file_sheets:
            return file_sheets[sheet_name], False, None
        
        # 回退检查默认表
        for default_sheet_name in self.default_fallback_sheets:
            ws = file_sheets.get(default_sheet_name)
            if ws and self._validate_sheet_headers(ws):
                Logger.debug(f"🛑 {sheet_name}不存在：检查回退表 [{default_sheet_name}]【可用】")
                return ws, True, default_sheet_name
        
        return None, None, None

    def _validate_sheet_headers(self, worksheet):
        """验证工作表表头是否包含关键字段"""
        try:
            # 读取首行作为表头
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            return all(field in header_row for field in self.key_fields)
        except Exception as e:
            Logger.error(f"表头验证失败: {str(e)}")
            return False
    def _get_folder_type(self, file_path):
        """从文件路径解析文件夹类型"""
        if os.path.commonpath([file_path, self.response_path]) == self.response_path:
            return os.path.basename(self.response_path)
        elif os.path.commonpath([file_path, self.snt_path]) == self.snt_path:
            return os.path.basename(self.snt_path)
        elif os.path.commonpath([file_path, self.report_path]) == self.report_path:
            return os.path.basename(self.report_path)
        else:
            return "other"

    # def _process_single_row(self, input_ws, fp, progress, snt_data, base_data, column_mapping):

    #     # 获取当前有效工作表的行生成器，检查REQUIRED_FIELDS是否存在数据，都不存在会报错
    #     data_gen = ExcelProcessor.excel_row_generator_skipping(
    #         input_ws,
    #         fp,
    #         progress,
    #         self.required_fields,
    #         strict_flag=False
    #     )
    #     # 如果一条数据也遍历不到，则当前的工作表无效——不存在任何REQUIRED_FIELDS有值的情况，回滚到默认表
    #     has_valid_data = False
    #     for row in data_gen:
    #         has_valid_data = True
    #         key = tuple(row[field] for field in self.key_fields)
    #         if key not in snt_data:
    #             Logger.debug(f"未找到匹配项: {key}，跳过更新")
    #             continue
            
    #         base_data[key].update(ExcelProcessor.column_mapping(row, column_mapping))
    #         # Logger.info(f"更新 {key} 的 {column_mapping} 列")
    #     return has_valid_data
    def _process_single_row(self, input_ws, fp, snt_data, base_data, column_mapping, data_lock=None):
        """处理单个工作表的行数据（线程安全版本）"""
        # 获取当前有效工作表的行生成器
        count = 0
        data_gen = ExcelProcessor.excel_row_generator_skipping(
            input_ws,
            fp,
            None,
            self.required_fields,
            strict_flag=False
        )
        
        has_valid_data = False
        for row in data_gen:
            has_valid_data = True
            # 如果不用字符串格式存储和读取，就会发生丢数据，匹配更新失败的情况！
            key = tuple(str(row[field]) for field in self.key_fields)
            if key not in snt_data:
                Logger.debug(f"未找到匹配项: {key}，跳过更新")
                continue
            
            # 使用锁保护共享数据的更新操作
            if data_lock:
                with data_lock:
                    base_data[key].update(ExcelProcessor.column_mapping(row, column_mapping))
            else:
                # 非并发场景下的原始逻辑
                base_data[key].update(ExcelProcessor.column_mapping(row, column_mapping))
            count += 1 

        Logger.debug(f"{fp} 更新 {count} 行数据")
        return has_valid_data
    
    def _process_single_file(self, sheets_wb_map, sheet_name, fp, snt_data, base_data, column_mapping, data_lock):
        """
        并发处理单个文件的数据，并安全地更新共享的 base_data 字典。

        参数:
        - sheets_wb_map (dict): self.sheet_maps[fp]，值为 Worksheet 对象。
        - sheet_name (str): 需要处理的目标工作表名称。
        - fp (str): 文件路径。
        - snt_data (dict): 基准数据（来自 SNT 文件），用于匹配关键字段。
        - base_data (dict): 共享字典，用于存储最终合并后的数据，key 为 key_fields 的元组。
        - column_mapping (dict): 列映射配置，用于将输入列与目标列对齐。
        - data_lock (threading.Lock): 线程锁对象，确保多线程环境下对 base_data 的安全访问。

        返回值:
        - None: 结果直接写入 base_data。

        异常处理:
        - 如果处理过程中发生错误，会记录日志但不会中断主线程。
        
        日志输出:
        - 如果找不到有效工作表或未找到有效数据，会记录警告信息。
        """
        try:
            # 获取有效工作表(如果找不到Sheet_name，则使用默认回退表)
            input_ws, is_defalut_sheet, rollback_sheet_name = self._get_valid_sheet(sheets_wb_map, sheet_name)
            if not input_ws:
                Logger.error(f"🛑 文件 {Path(fp).name} 无有效工作表")
                return
                
            # 调用原有的行处理方法（线程安全版本）
            roll_back = not self._process_single_row(input_ws, fp, snt_data, base_data, column_mapping, data_lock)
            
            # 若表中无数据，且使用的不是默认表，则尝试获取默认表数据
            if not is_defalut_sheet and roll_back:
                has_valid_data = False
                # 获取默认表
                for default_sheet_name in self.default_fallback_sheets:
                    input_ws, is_defalut_sheet, rollback_sheet_name = self._get_valid_sheet(sheets_wb_map, default_sheet_name)
                    # 默认表名有效——input_ws有值且is_defalut_sheet为false
                    if input_ws and not is_defalut_sheet and self._validate_sheet_headers(input_ws):
                        Logger.info(f"🛑 文件{fp}⏩ 使用回退表 [{default_sheet_name}]")
                        # 不短路
                        has_valid_data = has_valid_data | self._process_single_row(
                            input_ws, fp, snt_data, base_data, column_mapping, data_lock
                        )
                if not has_valid_data:
                    # 存在业务场景，sheet_name就是没有业务数据，也不存在默认表
                    Logger.info(f"⚠️ 文件{fp}:【{sheet_name}】中无有效数据")

            elif is_defalut_sheet and not roll_back:
                Logger.info(f"🛑 文件{fp}⏩ 使用回退表 [{rollback_sheet_name}]")

            Logger.info(f"✅ 文件{fp}⏩ 更新完成")
        except Exception as e:
            Logger.error(f"处理文件 {fp} 时发生错误: {str(e)}")
            raise RuntimeError ("测试")

    def _load_snt_data(self, sheet_name, headers):
        """
        将snt当前sheet_name数据存在关键字段keys——用于联系数据，的行写入内存{key_tuple,row}，并生成snt_map和fix_map映射后的结果数据base_data
        """
        try:
            snt_file = next((fp for fp in self.sheet_maps.keys() if self._get_folder_type(fp) == os.path.basename(self.snt_path)), None)
            if not snt_file:
                raise RuntimeError(f"未找到{self.snt_path}文件夹下的基准文件")

            snt_data = {}
            base_data = {}
            progress = ExcelProgressTracker()
            snt_ws = self.sheet_maps[snt_file][sheet_name]
            snt_gen = ExcelProcessor.excel_row_generator(
                snt_ws,
                snt_file,
                progress,
                self.key_fields,
                strict_flag=False
            )
            for row in snt_gen:
                # 如果不用字符串格式存储和读取，就会发生丢数据，匹配更新失败的情况！
                key = tuple(str(row[field]) for field in self.key_fields)
                if key in snt_data:
                    Logger.info(f"⚠️ 发现重复基准数据: {key}")
                snt_data[key] = row

            for key, snt_row in snt_data.items():
                # 获取目标列格式——也就是模板列格式
                base_row = {header: '' for header in headers}
                base_row.update(ExcelProcessor.fixed_mapping(self.fixed_mapping))
                base_row.update(ExcelProcessor.column_mapping(snt_row, self.snt_mapping))
                base_data[key] = base_row
                
            progress.close()
            Logger.info(f"📥 已加载 {len(snt_data)} 条有效基准数据")
            return snt_file, snt_data, base_data
        except Exception as e:
            raise RuntimeError (f"❌ 内存加载{snt_file}基准数据失败: {str(e)}")

    def _process_single_sheet(self, sheet_name, output_wb):
        """处理单个工作表"""
        try:
            # 获取当前sheet_name工作表的输出句柄、表头列表，用于后续处理
            output_ws = output_wb[sheet_name]
            headers = [cell.value for cell in output_ws[1]]
            Logger.info(f"{'='*75}")
            Logger.info(f"🔨 开始处理工作表 [{sheet_name}]")
            snt_file, snt_data, base_data = self._load_snt_data(sheet_name, headers)

            # 将sheet_maps——{fp_path:sheet_name:wb}中的fp按文件夹分类
            folder_sources = defaultdict(list)
            for fp in self.sheet_maps.keys():
                if fp == snt_file:
                    continue
                folder_sources[self._get_folder_type(fp)].append(fp)

            # 所有文件夹
            for folder, fps in folder_sources.items():
                Logger.info(f"🔄 正在处理 [{folder}] 文件夹内数据...")
                column_mapping = self.response_mapping if folder == 'res' else (self.report_mapping if folder == 'report' else None)
                # column_mapping =  self.response_mapping if folder == 'res' # TODO 扩充至report_mapping
                if not column_mapping:
                    raise RuntimeError (f"⚠️ 未找到 [{folder}] 的列映射配置")

                # 使用线程池并发处理文件
                data_lock = threading.Lock()
                with GlobalThreadPool.get_executor() as executor:
                    futures = [
                        executor.submit(
                            self._process_single_file, 
                            self.sheet_maps[fp], 
                            sheet_name, 
                            fp, 
                            snt_data, 
                            base_data, # 共享变量，线程安全
                            column_mapping, 
                            data_lock
                            ) 
                            for fp in fps
                            ]
                
                done, not_done = concurrent.futures.wait(futures, timeout = 60)

                # 等待所有任务完成
                for future in futures:
                    future.result()  # 获取结果，触发可能的异常
                    
            # ----------------------------
            # 阶段三：写入最终数据
            # ----------------------------
            # 排序按表头排序
            headers = [cell.value for cell in output_ws[1]]
            ordered_rows = ExcelProcessor.sort_generated_rows(base_data.values(), headers)
            list(map(lambda row: output_ws.append(row), ordered_rows))
            # 格式设置
            self._style_apply(output_ws)
            Logger.info(f"✅ 工作表 [{sheet_name}] 处理完成，共更新 {len(base_data.values())} 行数据")
            return True

        except Exception as e:
            Logger.error(f"❌ 工作表 [{sheet_name}] 处理失败: {str(e)}")
            Logger.debug(f"{traceback.format_exc()}")
            return False
    def run(self):
        """主执行流程"""
        try:
            # 阶段1：初始化配置
            self._load_mappings()
            self._validate_input_files()

            # 阶段2：准备输出文件——给用户反馈的snt文件
            absolute_path = FileProcessor.create_newfile_by_template(
                self.template_file,
                self.target_file,
                # 直接改模板文件就行
                # additional_columns=["列1", "列2"] 
            )
            output_wb = load_workbook(absolute_path)

            # 阶段3：多表处理
            success_flags = []
            for sheet_name in self.sheet_names:
                success_flags.append(
                    self._process_single_sheet(sheet_name, output_wb)
                )

            # 阶段4：保存结果，但凡有一个sheet处理失败，则删除不完整的输出文件
            if all(success_flags):
                output_wb.save(self.target_file)
                Logger.info(f"💾 结果文件保存成功: {self.target_file}")
                return True
            else:
                raise RuntimeError("部分工作表处理失败")
                
        except Exception as e:
            Logger.error(f"❌ 主流程执行失败: {str(e)}")
            Logger.debug(f"{traceback.format_exc()}")
            if os.path.exists(self.target_file):
                os.remove(self.target_file)
                Logger.error("已删除不完整的结果文件")
            return False
        finally:
            GlobalThreadPool.shutdown()

if __name__ == "__main__":
    processor = AutoSntProcessor()
    if processor.run():
        Logger.info("🎉 自动化处理完成！")
    else:
        Logger.error("❌ 处理过程中存在错误")
    input("按回车键退出...")