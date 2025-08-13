import streamlit as st
import pandas as pd
import os
from datetime import datetime
from pathlib import Path
from datetime import datetime
import plotly.express as px
from pathlib import Path
from openpyxl import load_workbook
from collections import Counter
import shutil
import io
from sinotrans.core import FileProcessor

# ==================== MODEL层 ====================
class ConfigLoader:
    """共享配置加载器"""
    DEFAULT_SHEET = "default_sheet"
    REQUIRED_SHEET = "required_sheet" 
    KEY_FIELDS = "key_fields"
    REQUIRED_FIELDS = "required_fields"
    
    def __init__(self, config_path="conf"):
        self.config_path = Path(config_path)
        self.sheet_config_file = self.config_path / "sheet_config.txt"
        self._load_config()
    
    def _load_config(self):
        """加载配置文件"""
        try:
            sheet_conf = FileProcessor.parse_mapping_dict(
                str(self.sheet_config_file), ':', '|', ',', '='
            )
            self.default_fallback_sheets = sheet_conf.get(self.DEFAULT_SHEET).field_name.split(",")
            self.key_fields = sheet_conf.get(self.KEY_FIELDS).field_name.split(",")
            self.required_fields = sheet_conf.get(self.REQUIRED_FIELDS).field_name.split(",")
            self.sheet_names = sheet_conf.get(self.REQUIRED_SHEET).field_name.split(",")
        except Exception as e:
            # 提供默认值
            self.key_fields = ['folder', 'po', 'lot']
            self.required_fields = ['fwd_feedback', 'Remark']
            self.sheet_names = ['CREATED', 'NOT INCLUDED', 'COORDINATED', 'REQUESTED', 'BOOKED']
            self.default_fallback_sheets = ['Sheet1', 'Follow UP']

class ConfigModel:
    """配置文件模型"""
    
    def __init__(self):
        self.config_path = Path("conf")
        self.config_files = {
            "sheet_config.txt": "工作表配置",
            "pending_po_mapping.txt": "PO映射配置", 
            "fixed_mapping.txt": "固定映射配置",
            "response_mapping.txt": "响应映射配置",
            "report_mapping.txt": "报告映射配置"
        }
    
    def get_config_files(self):
        """获取配置文件列表"""
        return self.config_files
    
    def read_config_file(self, filename):
        """读取配置文件内容"""
        try:
            file_path = self.config_path / filename
            if file_path.exists():
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read(), None
            else:
                return "", f"文件 {filename} 不存在"
        except Exception as e:
            return "", f"读取文件失败: {e}"
    
    def save_config_file(self, filename, content):
        """保存配置文件内容"""
        try:
            file_path = self.config_path / filename
            # 备份原文件
            if file_path.exists():
                backup_path = file_path.with_suffix(f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
                shutil.copy2(file_path, backup_path)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            return True, "保存成功"
        except Exception as e:
            return False, f"保存失败: {e}"

class FileUploadModel:
    """文件上传模型"""
    
    def __init__(self):
        self.upload_paths = {
            "SNT文件": Path("snt"),
            "响应文件": Path("res"), 
            "报告文件": Path("report")
        }
    
    def ensure_directories(self):
        """确保上传目录存在"""
        for path in self.upload_paths.values():
            path.mkdir(exist_ok=True)
    
    def save_uploaded_file(self, uploaded_file, file_type):
        """保存上传的文件"""
        try:
            self.ensure_directories()
            target_path = self.upload_paths[file_type]
            file_path = target_path / uploaded_file.name
            
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            return True, f"文件已保存到 {file_path}"
        except Exception as e:
            return False, f"保存文件失败: {e}"
    
    def get_uploaded_files(self):
        """获取已上传的文件列表"""
        files_info = {}
        for file_type, path in self.upload_paths.items():
            if path.exists():
                files = [f.name for f in path.glob("*.xlsx") if f.is_file()]
                files_info[file_type] = files
            else:
                files_info[file_type] = []
        return files_info
    
    def delete_file(self, file_type, filename):
        """删除指定文件"""
        try:
            file_path = self.upload_paths[file_type] / filename
            if file_path.exists():
                file_path.unlink()
                return True, f"文件 {filename} 已删除"
            else:
                return False, f"文件 {filename} 不存在"
        except Exception as e:
            return False, f"删除文件失败: {e}"

class DataModel:
    """数据模型 - 处理所有数据相关操作"""
    
    def __init__(self):
        self.config_path = Path("conf")
        self.target_path = Path("target")
        self.config_loader = ConfigLoader()
        self.required_fields = self.config_loader.required_fields
        self.key_fields = self.config_loader.key_fields
    
    def get_target_files(self):
        """获取target文件夹中的Excel文件"""
        if not self.target_path.exists():
            return []
        return [f for f in os.listdir(self.target_path) if f.endswith('.xlsx')]
    
    def get_sheet_names(self, file_path):
        """获取Excel文件中的工作表名称"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            return []
    
    def load_sheet_data(self, file_path, sheet_name):
        """加载指定工作表的数据"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df, None
        except Exception as e:
            return None, str(e)
    
    def analyze_sheet_data(self, df):
        """分析工作表数据"""
        analysis = {
            'total_records': len(df),
            'field_stats': {},
            'empty_records': [],
            'charts_data': {}
        }
        
        # 分析required_fields
        for field in self.required_fields:
            if field in df.columns:
                # 字段统计_排除Nan空值
                field_counts = df[field].value_counts(dropna=True)
                analysis['field_stats'][field] = {
                    'unique_count': df[field].nunique(),
                    'value_counts': field_counts.to_dict(),
                    'empty_count': df[field].isna().sum() + (df[field] == '').sum()
                }
                
                # 图表数据
                analysis['charts_data'][field] = field_counts
        
        # 查找异常数据（required_fields均为空）
        if all(field in df.columns for field in self.required_fields):
            empty_mask = df[self.required_fields].isna().all(axis=1) | (df[self.required_fields] == '').all(axis=1)
            analysis['empty_records'] = df[empty_mask].index.tolist()
        
        return analysis

class ProcessorModel:
    """处理器模型 - 处理SNT数据"""
    
    def __init__(self):
        self.target_path = Path("target")
    
    def process_data(self):
        """处理SNT数据"""
        try:
            # 清空目标文件夹
            # success, message = self.clear_target_folder()
            # if not success:
            #     return False, message
            
            # 导入并运行SNT2处理器
            from snt2 import AutoSntProcessor
            processor = AutoSntProcessor()
            result = processor.run()
            
            if result:
                return True, "数据处理成功完成！"
            else:
                return False, "数据处理失败，请检查日志"
        except Exception as e:
            return False, f"处理过程中出现错误: {e}"

# ==================== VIEW层 ====================
class BaseView:
    """基础视图类"""
    
    @staticmethod
    def show_header(title, subtitle=None):
        """显示页面标题"""
        st.markdown(f'''
        <div class="page-header">
            <h1 class="main-title">{title}</h1>
            {f'<p class="subtitle">{subtitle}</p>' if subtitle else ''}
        </div>
        ''', unsafe_allow_html=True)
    
    @staticmethod
    def show_success(message):
        """显示成功消息"""
        st.success(f"✅ {message}")
    
    @staticmethod
    def show_error(message):
        """显示错误消息"""
        st.error(f"❌ {message}")
    
    @staticmethod
    def show_info(message):
        """显示信息消息"""
        st.info(f"ℹ️ {message}")
    
    @staticmethod
    def create_card(title, content, icon="📋"):
        """创建卡片组件"""
        st.markdown(f'''
        <div class="custom-card">
            <div class="card-header">
                <span class="card-icon">{icon}</span>
                <h3 class="card-title">{title}</h3>
            </div>
            <div class="card-content">
                {content}
            </div>
        </div>
        ''', unsafe_allow_html=True)

class NavigationView(BaseView):
    """导航视图"""
    
    @staticmethod
    def render_navigation():
        """渲染导航栏"""
        # 应用标题
        st.markdown('''
        <div class="app-header">
            <h1 class="app-title">SNT数据处理工具</h1>
        </div>
        ''', unsafe_allow_html=True)
        
        # 导航按钮 - 改为4列布局
        st.markdown('<div class="nav-container">', unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            if st.button("⚙️ 配置管理", key="nav_config", use_container_width=True):
                st.session_state.current_page = '配置管理'
                st.rerun()
        
        with col2:
            if st.button("🚀 数据处理", key="nav_process", use_container_width=True):
                st.session_state.current_page = '数据处理'
                st.rerun()
        
        with col3:
            if st.button("📊 数据分析", key="nav_analysis", use_container_width=True):
                st.session_state.current_page = '数据分析'
                st.rerun()
        
        with col4:
            if st.button("📋 日志查看", key="nav_logs", use_container_width=True):
                st.session_state.current_page = '日志查看'
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # 初始化当前页面
        if 'current_page' not in st.session_state:
            st.session_state.current_page = '配置管理'

class ConfigView(BaseView):
    """配置管理视图"""
    
    @staticmethod
    def render(config_model):
        """渲染配置管理页面"""
        ConfigView.show_header("配置管理", "编辑和管理系统配置文件")
        
        # 配置文件选择
        config_files = config_model.get_config_files()
        
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.markdown('''
            <div class="config-sidebar">
                <h3>📁 配置文件</h3>
            </div>
            ''', unsafe_allow_html=True)
            
            selected_file = st.selectbox(
                "选择配置文件:",
                list(config_files.keys()),
                format_func=lambda x: f"{config_files[x]} ({x})"
            )
        
        with col2:
            if selected_file:
                st.markdown(f'''
                <div class="config-editor">
                    <h3>✏️ 编辑 {config_files[selected_file]}</h3>
                </div>
                ''', unsafe_allow_html=True)
                
                # 读取文件内容
                content, error = config_model.read_config_file(selected_file)
                
                if error:
                    ConfigView.show_error(error)
                else:
                    # 文件内容编辑器
                    edited_content = st.text_area(
                        "文件内容:",
                        value=content,
                        height=400,
                        help="编辑配置文件内容，保存前会自动备份原文件"
                    )
                    
                    # 保存按钮
                    if st.button("💾 保存配置", type="primary", use_container_width=True):
                        success, message = config_model.save_config_file(selected_file, edited_content)
                        if success:
                            ConfigView.show_success(message)
                            st.rerun()
                        else:
                            ConfigView.show_error(message)
                    
                    
                    # 配置文件说明
                    st.markdown('''
                    <div class="config-help">
                        <h4>📖 配置文件格式说明</h4>
                        <ul>
                            <li><strong>映射文件</strong>: 使用冒号(:)分隔键值对，每行一个映射关系</li>
                            <li><strong>工作表配置</strong>: 使用冒号(:)分隔配置项和值，用逗号分隔多个值</li>
                            <li><strong>示例</strong>: <code>source_field:target_field</code></li>
                        </ul>
                    </div>
                    ''', unsafe_allow_html=True)

class LogView(BaseView):
    """日志查看视图"""
    
    @staticmethod
    def render():
        """渲染日志查看页面"""
        LogView.show_header("日志查看", "查看系统运行日志和处理记录")
        
        try:
            logs_path = Path("logs")
            if not logs_path.exists():
                st.info("📝 日志文件夹不存在")
                return
                
            # 获取所有.log文件
            log_files = list(logs_path.glob("*.log"))
            
            if not log_files:
                st.info("📝 暂无日志文件")
                return
            
            # 日志文件选择和信息显示
            col1, col2 = st.columns([2, 1])
            
            with col1:
                # 日志文件选择
                selected_log = st.selectbox(
                    "📄 选择日志文件:",
                    log_files,
                    format_func=lambda x: f"{x.name} ({datetime.fromtimestamp(x.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')})",
                    index=len(log_files)-1 if log_files else 0  # 默认选择最新的
                )
            
            with col2:
                # 显示行数选择
                display_lines_count = st.selectbox(
                    "📊 显示行数:",
                    [100, 300, 500, 1000, "全部"],
                    index=1  # 默认300行
                )
            
            if selected_log:
                # 读取日志文件
                try:
                    with open(selected_log, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    
                    # 根据选择显示对应行数
                    if display_lines_count == "全部":
                        display_lines = lines
                    else:
                        display_lines = lines[-display_lines_count:] if len(lines) > display_lines_count else lines
                    
                    log_content = ''.join(display_lines)
                    
                    # 显示日志统计信息
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("📄 文件大小", f"{selected_log.stat().st_size / 1024:.1f} KB")
                    with col2:
                        st.metric("📊 总行数", len(lines))
                    with col3:
                        st.metric("👁️ 显示行数", len(display_lines))
                    with col4:
                        st.metric("🕒 修改时间", datetime.fromtimestamp(selected_log.stat().st_mtime).strftime('%H:%M:%S'))
                    
                    # 操作按钮
                    col1, col2, col3 = st.columns([1, 1, 1])
                    with col1:
                        if st.button("🔄 刷新日志", use_container_width=True, type="secondary"):
                            st.rerun()
                    
                    with col2:
                        # 下载日志文件
                        st.download_button(
                            label="📥 下载日志",
                            data=log_content,
                            file_name=selected_log.name,
                            mime="text/plain",
                            use_container_width=True,
                            type="secondary"
                        )
                    
                    with col3:
                        # 清空日志文件夹
                        if st.button("🗑️ 清空日志文件夹", use_container_width=True, type="secondary"):
                            # 确认对话框
                            if st.session_state.get('confirm_clear_logs', False):
                                try:
                                    logs_path = Path("logs")
                                    if logs_path.exists():
                                        # 删除logs文件夹内所有文件
                                        for file in logs_path.glob("*"):
                                            if file.is_file():
                                                file.unlink()
                                        st.success("✅ 日志文件夹已清空")
                                        st.session_state['confirm_clear_logs'] = False
                                        st.rerun()
                                    else:
                                        st.warning("⚠️ 日志文件夹不存在")
                                except Exception as e:
                                    st.error(f"❌ 清空日志失败: {str(e)}")
                                st.session_state['confirm_clear_logs'] = False
                            else:
                                st.session_state['confirm_clear_logs'] = True
                                st.warning("⚠️ 确认要清空所有日志文件吗？再次点击确认。")
                    
                    # 日志内容显示
                    st.markdown("### 📋 日志内容")
                    st.text_area(
                        "日志内容:",
                        value=log_content,
                        height=500,
                        help=f"显示 {selected_log.name} 的{'最后' if display_lines_count != '全部' else ''}{len(display_lines)} 行内容",
                        key="log_content_viewer"
                    )
                    
                    # 日志级别统计
                    if log_content:
                        LogView._show_log_statistics(log_content)
                        
                except Exception as e:
                    LogView.show_error(f"读取日志文件失败: {e}")
                    
        except Exception as e:
            LogView.show_error(f"获取日志文件失败: {e}")
    
    @staticmethod
    def _show_log_statistics(log_content):
        """显示日志统计信息"""
        st.markdown("### 📈 日志统计")
        
        # 统计不同级别的日志
        lines = log_content.split('\n')
        stats = {
            'INFO': sum(1 for line in lines if 'INFO' in line or '✅' in line or '📋' in line),
            'ERROR': sum(1 for line in lines if 'ERROR' in line or '❌' in line),
            'WARNING': sum(1 for line in lines if 'WARNING' in line or '⚠️' in line),
            'DEBUG': sum(1 for line in lines if 'DEBUG' in line)
        }
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("ℹ️ 信息", stats['INFO'])
        with col2:
            st.metric("❌ 错误", stats['ERROR'], delta_color="inverse")
        with col3:
            st.metric("⚠️ 警告", stats['WARNING'], delta_color="inverse")
        with col4:
            st.metric("🐛 调试", stats['DEBUG'])

class ProcessView(BaseView):
    """数据处理视图"""
    
    @staticmethod
    def render(controller):
        """渲染数据处理页面"""
        ProcessView.show_header("数据处理", "上传文件并处理SNT数据")
        
        # 创建两列布局
        col1, col2 = st.columns([1, 1])
        
        with col1:
            # 文件上传区域
            st.markdown('''
            <div class="upload-section">
                <h3>文件上传</h3>
            </div>
            ''', unsafe_allow_html=True)
            
            # 文件类型选择
            file_type = st.selectbox(
                "选择文件类型:",
                ["SNT文件", "响应文件", "报告文件"],
                help="选择要上传的文件类型"
            )
            
            # 文件上传
            uploaded_files = st.file_uploader(
                f"上传{file_type}:",
                type=['xlsx'],
                help="仅支持Excel文件(.xlsx格式)，可选择多个文件",
                accept_multiple_files=True
            )
            
            # 显示已选择的文件
            if uploaded_files:
                st.markdown("**已选择的文件:**")
                for i, file in enumerate(uploaded_files, 1):
                    st.write(f"{i}. {file.name}")
            
            if uploaded_files:
                if st.button(f"💾 上传文件 ({len(uploaded_files)}个)", type="primary", use_container_width=True):
                    success_count = 0
                    error_messages = []
                    
                    for uploaded_file in uploaded_files:
                        success, message = controller.file_upload_model.save_uploaded_file(uploaded_file, file_type)
                        if success:
                            success_count += 1
                        else:
                            error_messages.append(f"{uploaded_file.name}: {message}")
                    
                    if success_count == len(uploaded_files):
                        ProcessView.show_success(f"成功保存 {success_count} 个文件")
                        st.rerun()
                    elif success_count > 0:
                        ProcessView.show_success(f"成功保存 {success_count} 个文件")
                        for error in error_messages:
                            ProcessView.show_error(error)
                        st.rerun()
                    else:
                        for error in error_messages:
                            ProcessView.show_error(error)
                st.markdown("---")
            # 已上传文件列表
            st.markdown('''
            <div class="files-section">
                <h3>📁 已上传文件</h3>
            </div>
            ''', unsafe_allow_html=True)
            
            files_info = controller.file_upload_model.get_uploaded_files()
            
            for ftype, files in files_info.items():
                if files:
                    with st.expander(f"{ftype} ({len(files)}个文件)"):
                        for file in files:
                            col_file, col_delete = st.columns([3, 1])
                            with col_file:
                                st.text(file)
                            with col_delete:
                                if st.button("🗑️", key=f"delete_{ftype}_{file}", help="删除文件"):
                                    success, message = controller.file_upload_model.delete_file(ftype, file)
                                    if success:
                                        ProcessView.show_success(message)
                                        st.rerun()
                                    else:
                                        ProcessView.show_error(message)
        
        with col2:
            # 数据处理区域
            st.markdown('''
            <div class="process-section">
                <h3>数据处理</h3>
            </div>
            ''', unsafe_allow_html=True)
            
            # 处理状态显示
            if 'process_result' in st.session_state:
                result = st.session_state.process_result
                if result['success']:
                    st.success(f"✅ {result['message']}")
                else:
                    st.error(f"❌ {result['message']}")
            
            # 处理按钮
            if st.button("开始处理", type="primary", use_container_width=True, help="开始处理SNT数据"):
                controller.process_data()
            
            # 处理结果下载
            if 'process_result' in st.session_state and st.session_state.process_result['success']:
                target_files = controller.data_model.get_target_files()
                if target_files:
                    st.markdown('''
                    <div class="download-section">
                        <h4>下载结果</h4>
                    </div>
                    ''', unsafe_allow_html=True)
                    
                    latest_file = max(target_files, key=lambda x: os.path.getctime(os.path.join("target", x)))
                    file_path = os.path.join("target", latest_file)
                    
                    with open(file_path, 'rb') as f:
                        st.download_button(
                            label=f"📥 下载 {latest_file}",
                            data=f.read(),
                            file_name=latest_file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
            
            # 处理说明
            st.markdown('''
            <div class="process-help">
                <h4>处理说明</h4>
                <ul>
                    <li>确保已上传所需的SNT、响应和报告文件</li>
                    <li>处理完成后可下载生成的结果文件</li>
                    <li>如遇问题请检查日志文件</li>
                </ul>
            </div>
            ''', unsafe_allow_html=True)

class AnalysisView(BaseView):
    """数据分析视图"""
    
    @staticmethod
    def render(controller):
        """渲染数据分析页面"""
        AnalysisView.show_header("数据分析", "分析处理结果和数据质量")
        
        # 文件选择区域
        target_files = controller.data_model.get_target_files()
        
        if not target_files:
            st.markdown('''
            <div class="no-data-message">
                <h3>📂 暂无数据文件</h3>
                <p>请先在'数据处理'页面生成数据文件</p>
            </div>
            ''', unsafe_allow_html=True)
            return
        
        # 文件和工作表选择
        col1, col2 = st.columns(2)
        
        with col1:
            selected_file = st.selectbox("📁 选择分析文件:", target_files)
        
        with col2:
            if selected_file:
                file_path = os.path.join("target", selected_file)
                sheet_names = controller.data_model.get_sheet_names(file_path)
                
                if sheet_names:
                    selected_sheet = st.selectbox("📋 选择工作表:", sheet_names)
                else:
                    AnalysisView.show_error("无法读取文件中的工作表")
                    return
        
        if selected_file and selected_sheet:
            # 加载和分析数据
            df, error = controller.data_model.load_sheet_data(file_path, selected_sheet)
            
            if error:
                AnalysisView.show_error(f"加载数据失败: {error}")
                return
            
            analysis = controller.data_model.analyze_sheet_data(df)
            
            # 数据概览卡片
            # st.markdown('''
            # <div class="metrics-container">
            #     <h3>数据概览</h3>
            # </div>
            # ''', unsafe_allow_html=True)
            st.markdown(f'''
                <div class="chart-section">
                    <h3>数据概览</h3>
                </div>
                ''', unsafe_allow_html=True)
            # 创建指标卡片
            cols = st.columns(len(controller.data_model.required_fields) + 2)
            
            with cols[0]:
                st.metric(
                    label="📈 总记录数",
                    value=analysis['total_records'],
                    help="工作表中的总记录数量"
                )
            
            with cols[1]:
                empty_count = len(analysis['empty_records'])
                st.metric(
                    label="⚠️ 异常记录",
                    value=empty_count,
                    # delta=f"{empty_count/analysis['total_records']*100:.1f}%" if analysis['total_records'] > 0 else "0%",
                    # delta_color="inverse",
                    help="required_fields字段均为空的记录数"
                )
            
            for i, field in enumerate(controller.data_model.required_fields):
                if field in analysis['field_stats']:
                    with cols[i + 2]:
                        stats = analysis['field_stats'][field]
                        st.metric(
                            label=f"🏷️ {field}类型",
                            value=stats['unique_count'],
                            help=f"{field}字段的唯一值数量"
                        )
            
            # 字段分析图表
            for field in controller.data_model.required_fields:
                if field in analysis['charts_data']:
                    st.markdown(f'''
                    <div class="chart-section">
                        <h3>{field} 分布分析</h3>
                    </div>
                    ''', unsafe_allow_html=True)
                    
                    chart_data = analysis['charts_data'][field]
                    if len(chart_data) > 0:
                        # 创建图表
                        chart_df = pd.DataFrame({
                            'category': chart_data.index,
                            'count': chart_data.values
                        })
                        
                        fig = px.bar(
                            chart_df, 
                            x='count', 
                            y='category',
                            orientation='h',
                            title=f"{field} 类型分布",
                            color='count',
                            color_continuous_scale='viridis'
                        )
                        fig.update_layout(
                            height=400,
                            showlegend=False,
                            title_x=0.5
                        )
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # 显示统计信息
                        stats = analysis['field_stats'][field]
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.info(f"🏆 最常见类型: {chart_data.index[0]}")
                        with col2:
                            st.info(f"📊 出现次数: {chart_data.iloc[0]}")
                        with col3:
                            st.info(f"❌ 空值数量: {stats['empty_count']}")
            
            # 详细数据表
            with st.expander("📋 查看详细数据"):
                st.dataframe(df, use_container_width=True)
                # 导出功能
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='分析数据', index=False)
                excel_data = excel_buffer.getvalue()
                
                st.download_button(
                    label="📤 导出xlsx数据",
                    data=excel_data,
                    file_name=f"{selected_sheet}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            # 异常数据展示
            if analysis['empty_records']:
                st.markdown('''
                <div class="alert-section">
                    <h3>⚠️ 异常数据检测</h3>
                </div>
                ''', unsafe_allow_html=True)
                
                st.error(f"发现 {len(analysis['empty_records'])} 条记录的required_fields字段均为空")
                
                with st.expander("🔍 查看异常记录详情"):
                    empty_df = df.iloc[analysis['empty_records']]
                    st.dataframe(empty_df, use_container_width=True)
                    
                    # 导出异常数据
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        empty_df.to_excel(writer, sheet_name='异常数据', index=False)
                    excel_data = excel_buffer.getvalue()
                    
                    st.download_button(
                        label="📤 导出异常数据",
                        data=excel_data,
                        file_name=f"异常数据_{selected_sheet}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.success("✅ 未发现异常数据，数据质量良好")

# ==================== CONTROLLER层 ====================
class MainController:
    """主控制器"""
    
    def __init__(self):
        self.data_model = DataModel()
        self.processor_model = ProcessorModel()
        self.config_model = ConfigModel()
        self.file_upload_model = FileUploadModel()
    
    def process_data(self):
        """处理数据"""
        with st.spinner("正在处理数据，请稍候..."):
            success, message = self.processor_model.process_data()
            st.session_state.process_result = {
                'success': success,
                'message': message
            }
            st.rerun()

# ==================== 主应用 ====================
def setup_page_config():
    """设置页面配置"""
    st.set_page_config(
        page_title="SNT数据处理工具",
        # page_icon="🚀",
        layout="wide",
        initial_sidebar_state="collapsed"
    )

def setup_custom_css():
    """设置自定义CSS样式"""
    st.markdown("""
    <style>
    /* 全局样式 */
    .main {
        padding-top: 1rem;
    }
    
    /* 应用头部 */
    .app-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 8px 32px rgba(0,0,0,0.1);
    }
    
    .app-header .app-title,
    .app-header h1.app-title,
    .app-title {
        color: white !important;
        font-size: 2.5rem !important;
        font-weight: bold !important;
        margin: 0 !important;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3) !important;
    }
    
    /* 页面头部 */
    .page-header {
        background: linear-gradient(90deg, #f8f9fa, #e9ecef);
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        border-left: 5px solid #007bff;
    }
    
    .main-title {
        color: #2c3e50;
        font-size: 2.5rem;
        font-weight: bold;
        margin: 0;
    }
    
    .subtitle {
        color: #6c757d;
        font-size: 1.1rem;
        margin: 0.5rem 0 0 0;
    }
    
    /* 导航容器 */
    .nav-container {
        margin: 2rem 0;
    }
    
    .page-indicator {
        text-align: center;
        margin: 1rem 0;
        padding: 0.5rem;
        background: #f8f9fa;
        border-radius: 20px;
    }
    
    .current-page {
        color: #007bff;
        font-weight: bold;
    }
    
    /* 按钮样式 */
    .stButton > button {
        width: 100%;
        border-radius: 25px;
        font-weight: bold;
        transition: all 0.3s ease;
        border: none;
        padding: 0.75rem 1.5rem;
        font-size: 1rem;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
    
    /* 卡片样式 */
    .custom-card {
        background: white;
        border-radius: 15px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        border: 1px solid #e9ecef;
    }
    
    .card-header {
        display: flex;
        align-items: center;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #f8f9fa;
    }
    
    .card-icon {
        font-size: 1.5rem;
        margin-right: 0.5rem;
    }
    
    .card-title {
        color: #2c3e50;
        margin: 0;
        font-size: 1.3rem;
    }
    
    /* 配置页面样式 */
    .config-sidebar {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
    }
    
    .config-editor {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        border: 1px solid #dee2e6;
    }
    
    .config-help {
        background: #e7f3ff;
        padding: 1rem;
        border-radius: 10px;
        margin-top: 1rem;
        border-left: 4px solid #007bff;
    }
    
    /* 上传区域样式 */
    .upload-section {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        text-align: center;
    }
    
    .files-section {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .process-section {
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        text-align: center;
    }
    
    .download-section {
        background: #e8f5e8;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
        border-left: 4px solid #28a745;
    }
    
    .process-help {
        background: #fff3cd;
        padding: 1rem;
        border-radius: 10px;
        margin-top: 1rem;
        border-left: 4px solid #ffc107;
    }
    
    /* 日志页面样式 */
    .log-info-section {
        margin: 1rem 0;
    }
    
    .log-section {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        text-align: center;
    }
    
    /* 分析页面样式 */
    .metrics-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        text-align: center;
    }
    
    .alert-section {
        background: #f8d7da;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
        border-left: 4px solid #dc3545;
    }
    
    .chart-section {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
        border-left: 4px solid #17a2b8;
    }
    
    .no-data-message {
        text-align: center;
        padding: 3rem;
        background: #f8f9fa;
        border-radius: 15px;
        margin: 2rem 0;
    }
    
    /* 指标卡片 */
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        text-align: center;
        margin: 0.5rem;
        border-top: 4px solid #007bff;
    }
    
    /* 响应式设计 */
    @media (max-width: 768px) {
        .app-title {
            font-size: 2rem;
        }
        
        .main-title {
            font-size: 1.8rem;
        }
        
        .custom-card {
            margin: 0.5rem 0;
            padding: 1rem;
        }
    }
    </style>
    """, unsafe_allow_html=True)

def main():
    """主函数 - MVC架构入口"""
    # 页面配置
    setup_page_config()
    setup_custom_css()
    
    # 创建控制器
    controller = MainController()
    
    # 渲染导航
    NavigationView.render_navigation()
    
    # 根据当前页面渲染对应视图
    current_page = st.session_state.get('current_page', '配置管理')
    
    if current_page == '配置管理':
        ConfigView.render(controller.config_model)
    elif current_page == '数据处理':
        ProcessView.render(controller)
    elif current_page == '数据分析':
        AnalysisView.render(controller)
    elif current_page == '日志查看':
        LogView.render()
    
    # 页脚
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: #666; padding: 1rem;'>"
        "SNT数据处理工具 v3.0 | Created by qiuyutian | Powered by Streamlit"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()